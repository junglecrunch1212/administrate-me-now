"""
demo_xlsx_roundtrip.py — end-to-end smoke for phase 07c-β.

Spins up a tmp instance, registers projections + forward + reverse
daemons, appends ~10 seed events, regenerates both workbooks, performs a
programmatic xlsx edit on each, runs reverse cycles, regenerates again,
and prints event counts at each stage. Exits 0 in well under 30s.

Invocation: ``poetry run python scripts/demo_xlsx_roundtrip.py``.
"""

from __future__ import annotations

import asyncio
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.reverse import XlsxReverseDaemon
from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

KEY = b"d" * 32


def _env(event_type: str, payload: dict) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="demo-tenant",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="demo",
        source_account_id="demo",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise RuntimeError(f"{sid} lagged: {s}")


async def _count_events(log: EventLog) -> dict[str, int]:
    counts: dict[str, int] = {}
    async for e in log.read_since(None):
        counts[e["type"]] = counts.get(e["type"], 0) + 1
    return counts


async def main() -> None:
    with tempfile.TemporaryDirectory() as td_str:
        td = Path(td_str)
        (td / "config").mkdir()
        (td / "config" / "instance.yaml").write_text("tenant_id: demo-tenant\n")
        config = load_instance_config(td)
        log = EventLog(config, KEY)
        bus = EventBus(log, config.bus_checkpoint_path)
        runner = ProjectionRunner(bus, log, config, encryption_key=KEY)
        runner.register(PartiesProjection())
        runner.register(TasksProjection())
        runner.register(CommitmentsProjection())
        runner.register(RecurrencesProjection())
        runner.register(CalendarsProjection())
        runner.register(PlacesAssetsAccountsProjection())
        runner.register(MoneyProjection())
        await runner.start()
        try:
            for i in range(2):
                await log.append(_env("party.created", {
                    "party_id": f"p{i}", "kind": "person",
                    "display_name": f"P{i}", "sort_name": f"P{i}",
                }))
            for i in range(3):
                await log.append(_env("task.created", {
                    "task_id": f"t{i}", "title": f"Task {i}",
                }))
            await log.append(_env("recurrence.added", {
                "recurrence_id": "r1",
                "linked_kind": "household",
                "linked_id": "household",
                "kind": "maintenance",
                "rrule": "FREQ=MONTHLY",
                "next_occurrence": "2026-05-01T09:00:00Z",
            }))
            await log.append(_env("account.added", {
                "account_id": "a1", "display_name": "Checking",
                "organization_party_id": "p0", "kind": "bank",
                "status": "active", "attributes": {"last4": "1234"},
            }))
            await log.append(_env("money_flow.recorded", {
                "flow_id": "fp1", "amount_minor": 1000, "currency": "USD",
                "occurred_at": "2026-04-20T10:00:00Z",
                "kind": "paid", "linked_account": "a1",
                "source_adapter": "plaid",
            }))
            last = await log.append(_env("money_flow.manually_added", {
                "flow_id": "fm1", "amount_minor": 4321, "currency": "USD",
                "occurred_at": "2026-04-22T10:00:00Z",
                "kind": "paid",
                "category": "groceries",
                "added_by_party_id": "p0",
            }))
            assert last is not None
            await bus.notify(last)
            for name in (
                "parties", "tasks", "commitments", "recurrences",
                "calendars", "places_assets_accounts", "money",
            ):
                await _wait_idle(bus, f"projection:{name}")

            counts_seed = await _count_events(log)
            print(f"[stage 1] seed events appended: {sum(counts_seed.values())}")

            ctx = XlsxQueryContext(
                parties_conn=runner.connection("parties"),
                tasks_conn=runner.connection("tasks"),
                commitments_conn=runner.connection("commitments"),
                recurrences_conn=runner.connection("recurrences"),
                calendars_conn=runner.connection("calendars"),
                places_assets_accounts_conn=runner.connection(
                    "places_assets_accounts"
                ),
                money_conn=runner.connection("money"),
            )
            forward = XlsxWorkbooksProjection(
                config, ctx, event_log=log, debounce_s=0.05
            )
            await forward.regenerate_now(OPS_WORKBOOK_NAME)
            await forward.regenerate_now(FINANCE_WORKBOOK_NAME)
            counts_post_forward = await _count_events(log)
            print(
                f"[stage 2] xlsx.regenerated count: "
                f"{counts_post_forward.get('xlsx.regenerated', 0)}"
            )

            # Programmatic principal-style edit: rename a Tasks row,
            # append a manual money flow.
            ops_path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
            wb = load_workbook(str(ops_path))
            ws = wb["Tasks"]
            ws.cell(row=2, column=2, value="Renamed Task 0")
            wb.save(str(ops_path))
            wb.close()

            finance_path = config.xlsx_workbooks_dir / FINANCE_WORKBOOK_NAME
            wb = load_workbook(str(finance_path))
            ws = wb["Raw Data"]
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value="")  # mint
            ws.cell(row=new_row, column=2, value="2026-04-25")
            ws.cell(row=new_row, column=6, value=15.50)
            ws.cell(row=new_row, column=9, value="snacks")
            ws.cell(row=new_row, column=11, value=True)
            wb.save(str(finance_path))
            wb.close()

            reverse = XlsxReverseDaemon(
                config,
                ctx,
                event_log=log,
                flush_wait_s=0.05,
                forward_lock_timeout_s=2.0,
                delete_undo_window_s=0.1,
            )
            await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
            await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)
            await asyncio.sleep(0.3)

            counts_post_reverse = await _count_events(log)
            print(
                f"[stage 3] task.updated: "
                f"{counts_post_reverse.get('task.updated', 0)}, "
                f"money_flow.manually_added: "
                f"{counts_post_reverse.get('money_flow.manually_added', 0)}, "
                f"xlsx.reverse_projected: "
                f"{counts_post_reverse.get('xlsx.reverse_projected', 0)}, "
                f"xlsx.reverse_skipped_during_forward: "
                f"{counts_post_reverse.get('xlsx.reverse_skipped_during_forward', 0)}"
            )

            # Notify the bus so the projections see the reverse-emitted
            # domain events, then re-regenerate forward.
            latest = await log.latest_event_id()
            assert latest is not None
            await bus.notify(latest)
            for name in ("tasks", "money"):
                await _wait_idle(bus, f"projection:{name}")
            await forward.regenerate_now(OPS_WORKBOOK_NAME)
            await forward.regenerate_now(FINANCE_WORKBOOK_NAME)
            counts_final = await _count_events(log)
            print(
                f"[stage 4] final xlsx.regenerated: "
                f"{counts_final.get('xlsx.regenerated', 0)}, "
                f"total events: {sum(counts_final.values())}"
            )

            await reverse.stop()
        finally:
            await runner.stop()
            await log.close()


if __name__ == "__main__":
    asyncio.run(main())
