"""
demo_xlsx_forward.py — end-to-end smoke for phase 07b.

Spins up a tmp instance, registers the 10 projections plus
XlsxWorkbooksProjection, appends a small scripted event set, regenerates
both workbooks synchronously (bypassing the 5s debounce), and prints the
resulting file paths with per-sheet row counts.

Runs in under 30s. No user interaction. Exits 0 on success.

Invocation: ``poetry run python scripts/demo_xlsx_forward.py``.
"""

from __future__ import annotations

import asyncio
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook

from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.artifacts import ArtifactsProjection
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.interactions import InteractionsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.vector_search import VectorSearchProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

KEY = b"s" * 32


def _env(event_type: str, payload: dict) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="demo-tenant",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="demo",
        source_account_id="demo",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise RuntimeError(f"{sid} lagged: {s}")


async def main() -> None:
    with tempfile.TemporaryDirectory() as td_str:
        td = Path(td_str)
        (td / "config").mkdir()
        (td / "config" / "instance.yaml").write_text("tenant_id: demo-tenant\n")
        config = load_instance_config(td)
        log = EventLog(config, KEY)
        bus = EventBus(log, config.bus_checkpoint_path)
        runner = ProjectionRunner(bus, log, config, encryption_key=KEY)
        runner.register(PartiesProjection())
        runner.register(InteractionsProjection())
        runner.register(ArtifactsProjection())
        runner.register(CommitmentsProjection())
        runner.register(TasksProjection())
        runner.register(RecurrencesProjection())
        runner.register(CalendarsProjection())
        runner.register(PlacesAssetsAccountsProjection())
        runner.register(MoneyProjection())
        runner.register(VectorSearchProjection())
        await runner.start()

        try:
            # Tiny but varied fixture.
            for i in range(3):
                await log.append(_env("party.created", {
                    "party_id": f"p{i}", "kind": "person",
                    "display_name": f"P{i}", "sort_name": f"P{i}",
                }))
            for i in range(4):
                await log.append(_env("task.created", {
                    "task_id": f"t{i}", "title": f"Task {i}",
                    "owner_member_id": f"p{i % 3}",
                }))
            await log.append(_env("recurrence.added", {
                "recurrence_id": "r1",
                "linked_kind": "household",
                "linked_id": "hh",
                "kind": "maintenance",
                "rrule": "FREQ=MONTHLY",
                "next_occurrence": "2026-05-01T09:00:00Z",
            }))
            await log.append(_env("account.added", {
                "account_id": "a1", "display_name": "Checking",
                "organization_party_id": "p0", "kind": "bank",
                "status": "active", "attributes": {"last4": "1234"},
            }))
            await log.append(_env("money_flow.recorded", {
                "flow_id": "f1", "amount_minor": 2500, "currency": "USD",
                "occurred_at": "2026-04-22T10:00:00Z",
                "kind": "paid", "linked_account": "a1",
                "source_adapter": "plaid",
            }))

            latest = await log.latest_event_id()
            assert latest is not None
            await bus.notify(latest)
            for name in (
                "parties", "tasks", "commitments", "recurrences",
                "calendars", "places_assets_accounts", "money",
            ):
                await _wait_idle(bus, f"projection:{name}")

            ctx = XlsxQueryContext(
                parties_conn=runner.connection("parties"),
                tasks_conn=runner.connection("tasks"),
                commitments_conn=runner.connection("commitments"),
                recurrences_conn=runner.connection("recurrences"),
                calendars_conn=runner.connection("calendars"),
                places_assets_accounts_conn=runner.connection(
                    "places_assets_accounts"
                ),
                money_conn=runner.connection("money"),
            )
            projection = XlsxWorkbooksProjection(
                config, ctx, event_log=log, debounce_s=0.1
            )
            await projection.regenerate_now(OPS_WORKBOOK_NAME)
            await projection.regenerate_now(FINANCE_WORKBOOK_NAME)

            ops_path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
            finance_path = config.xlsx_workbooks_dir / FINANCE_WORKBOOK_NAME

            print(f"[smoke] ops workbook: {ops_path}")
            wb_ops = load_workbook(str(ops_path))
            for name in wb_ops.sheetnames:
                ws = wb_ops[name]
                print(f"  sheet {name}: {ws.max_row - 1} data rows")

            print(f"[smoke] finance workbook: {finance_path}")
            wb_fin = load_workbook(str(finance_path))
            for name in wb_fin.sheetnames:
                ws = wb_fin[name]
                print(f"  sheet {name}: {ws.max_row - 1} data rows")

            # Count xlsx.regenerated events
            count = 0
            async for e in log.read_since(None):
                if e["type"] == "xlsx.regenerated":
                    count += 1
            print(f"[smoke] xlsx.regenerated events emitted: {count}")

        finally:
            await runner.stop()
            await log.close()


if __name__ == "__main__":
    asyncio.run(main())
