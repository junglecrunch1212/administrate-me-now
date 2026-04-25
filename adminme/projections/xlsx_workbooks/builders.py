"""
Workbook-level builders — assemble sheets, write atomically.

Per ADMINISTRATEME_BUILD.md §3.11 forward-projection algorithm step 7:
write to a temp file, rename atomically. The lock is acquired and
released by the caller (the projection's ``_regenerate`` coroutine).

Per prompt 08a: each workbook builder constructs a single internal
Session via ``build_internal_session("xlsx_workbooks", "device", tenant_id)``
at entry [§6.1, BUILD.md L3-continued]. The session is passed to each
sheet builder; sheets that route reads through the projection
``queries.py`` modules (a future refactor) will use it directly. The
current sheet builders read via raw SQL on the projection connections —
they accept the session for forward-compatibility and to anchor the
session-construction site in one place per workbook.
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook

from adminme.lib.session import build_internal_session
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext
from adminme.projections.xlsx_workbooks.sheets import (
    accounts as accounts_sheet,
    commitments as commitments_sheet,
    metadata_finance,
    metadata_ops,
    people as people_sheet,
    raw_data as raw_data_sheet,
    recurrences as recurrences_sheet,
    tasks as tasks_sheet,
)


def build_ops_workbook(
    path: Path,
    ctx: XlsxQueryContext,
    *,
    tenant_id: str,
    last_event_id: str,
) -> None:
    """Regenerate adminme-ops.xlsx atomically at ``path``."""
    path.parent.mkdir(parents=True, exist_ok=True)
    # Internal session anchored once per workbook build per [§6.1]; sheet
    # builders accept and propagate it for any query-function calls.
    session = build_internal_session("xlsx_workbooks", "device", tenant_id)
    wb = Workbook()
    # Remove the default sheet openpyxl creates.
    default = wb.active
    if default is not None:
        wb.remove(default)

    tasks_ws = wb.create_sheet("Tasks")
    tasks_sheet.build(tasks_ws, ctx, tenant_id=tenant_id, session=session)

    recurrences_ws = wb.create_sheet("Recurrences")
    recurrences_sheet.build(
        recurrences_ws, ctx, tenant_id=tenant_id, session=session
    )

    commitments_ws = wb.create_sheet("Commitments")
    commitments_sheet.build(
        commitments_ws, ctx, tenant_id=tenant_id, session=session
    )

    people_ws = wb.create_sheet("People")
    people_sheet.build(people_ws, ctx, tenant_id=tenant_id, session=session)

    metadata_ws = wb.create_sheet("Metadata")
    metadata_ops.build(
        metadata_ws,
        ctx,
        tenant_id=tenant_id,
        last_event_id=last_event_id,
        session=session,
    )

    tmp = path.with_suffix(path.suffix + f".tmp.{os.getpid()}")
    try:
        wb.save(str(tmp))
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def build_finance_workbook(
    path: Path,
    ctx: XlsxQueryContext,
    *,
    tenant_id: str,
    last_event_id: str,
) -> None:
    """Regenerate adminme-finance.xlsx atomically at ``path``.

    Sheets built: Raw Data, Accounts, Metadata. Per the 07b prompt,
    Assumptions / Dashboard / Balance Sheet / 5-Year Pro Forma / Budget
    vs Actual are deferred — they depend on unregistered event types or
    derived-math pipelines landing in prompt 10c+.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    session = build_internal_session("xlsx_workbooks", "device", tenant_id)
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    raw_ws = wb.create_sheet("Raw Data")
    raw_data_sheet.build(raw_ws, ctx, tenant_id=tenant_id, session=session)

    accounts_ws = wb.create_sheet("Accounts")
    accounts_sheet.build(accounts_ws, ctx, tenant_id=tenant_id, session=session)

    metadata_ws = wb.create_sheet("Metadata")
    metadata_finance.build(
        metadata_ws,
        ctx,
        tenant_id=tenant_id,
        last_event_id=last_event_id,
        session=session,
    )

    tmp = path.with_suffix(path.suffix + f".tmp.{os.getpid()}")
    try:
        wb.save(str(tmp))
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
