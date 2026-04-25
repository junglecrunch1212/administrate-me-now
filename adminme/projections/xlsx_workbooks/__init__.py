"""
xlsx_workbooks — forward-only xlsx projection daemon (prompt 07b).

xlsx_workbooks is structurally a projection per [§2.2]: it consumes
events and regenerates derived state. It emits exactly one event,
``xlsx.regenerated``, which is categorized as a SYSTEM event (not a
domain event). System events are observability-only and do not
represent facts about the world. This resolves the apparent tension
between BUILD.md §3.11 step 9 ("emit xlsx.regenerated") and [§2.2]
("projections never emit"). Prompt 07.5's audit relies on this
classification.

Unlike other L3 projections, this one does NOT own a SQLite backing
store: the derived state is two files on disk (``adminme-ops.xlsx`` and
``adminme-finance.xlsx``) under ``InstanceConfig.xlsx_workbooks_dir``.
``apply()`` does not write DB rows; instead it schedules a debounced
regeneration per workbook.

Sheets built in this prompt:
- adminme-ops.xlsx: Tasks, Recurrences, Commitments, People, Metadata.
- adminme-finance.xlsx: Raw Data, Accounts, Metadata (prompt 07b-3).

Sheets deferred to future prompts (await unregistered event types or
derived math):
- Lists, Members, Assumptions, Dashboard, Balance Sheet, 5-Year Pro
  Forma, Budget vs Actual.

Trigger events are limited to the 40 types registered as of phase 07a.
TODO(future-prompt): extend subscription when these event types register:
  list_item.added, list_item.completed, list_item.removed,
  member.created, member.profile_changed, member.role_changed,
  party.tag.added, party.tag.removed, party.tier.computed,
  assumption.added, assumption.updated, assumption.removed,
  plaid.sync.completed, plaid.go_live,
  institution.added, institution.updated.

Construction is different from other projections because xlsx reads
across seven projection databases. The runner's register() still works
unchanged; the bootstrap (or test fixture) constructs
XlsxWorkbooksProjection(config, query_context, event_log=log) with
connection handles gathered after the other projections have started.
"""

from __future__ import annotations

import asyncio
import logging
import time
from pathlib import Path
from typing import Any

from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import InstanceConfig
from adminme.projections.base import Projection
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext
from adminme.projections.xlsx_workbooks.sidecar import (
    write_readonly_state,
    write_sheet_state,
)

_log = logging.getLogger(__name__)

OPS_WORKBOOK_NAME = "adminme-ops.xlsx"
FINANCE_WORKBOOK_NAME = "adminme-finance.xlsx"

# Single source of truth for which sheets each workbook contains and
# whether reverse-projection treats them as bidirectional or read-only.
# Drives both the ``xlsx.regenerated`` payload's ``sheets_regenerated``
# list and the per-cycle sidecar writer.
_BIDIRECTIONAL_SHEETS: dict[str, list[str]] = {
    OPS_WORKBOOK_NAME: ["Tasks", "Recurrences", "Commitments"],
    FINANCE_WORKBOOK_NAME: ["Raw Data"],
}
_READONLY_SHEETS: dict[str, list[str]] = {
    OPS_WORKBOOK_NAME: ["People", "Metadata"],
    FINANCE_WORKBOOK_NAME: ["Accounts", "Metadata"],
}

# adminme-ops.xlsx trigger event types (currently registered only).
_OPS_TRIGGERS: tuple[str, ...] = (
    "task.created",
    "task.updated",
    "task.completed",
    "task.deleted",
    "commitment.proposed",
    "commitment.confirmed",
    "commitment.completed",
    "commitment.dismissed",
    "commitment.edited",
    "commitment.snoozed",
    "commitment.cancelled",
    "commitment.delegated",
    "commitment.expired",
    "recurrence.added",
    "recurrence.updated",
    "recurrence.completed",
    "party.created",
    "party.merged",
    "identifier.added",
    "membership.added",
    "relationship.added",
    "calendar.event_added",
    "calendar.event_updated",
    "calendar.event_deleted",
)

# adminme-finance.xlsx trigger event types (currently registered only).
_FINANCE_TRIGGERS: tuple[str, ...] = (
    "money_flow.recorded",
    "money_flow.manually_added",
    "money_flow.manually_deleted",
    "account.added",
    "account.updated",
    "place.added",
    "place.updated",
    "asset.added",
    "asset.updated",
)

_TYPE_TO_WORKBOOK: dict[str, str] = {}
for _t in _OPS_TRIGGERS:
    _TYPE_TO_WORKBOOK[_t] = OPS_WORKBOOK_NAME
for _t in _FINANCE_TRIGGERS:
    _TYPE_TO_WORKBOOK[_t] = FINANCE_WORKBOOK_NAME


class XlsxWorkbooksProjection(Projection):
    """Forward-only xlsx projection daemon.

    Structurally a projection per [§2.2]. Emits only the system event
    ``xlsx.regenerated`` after each successful regenerate.
    """

    name = "xlsx_workbooks"
    version = 1
    subscribes_to: list[str] = list(_OPS_TRIGGERS) + list(_FINANCE_TRIGGERS)
    schema_path: Path = Path(__file__).parent / "schema.sql"

    def __init__(
        self,
        config: InstanceConfig,
        query_context: XlsxQueryContext,
        *,
        event_log: EventLog,
        debounce_s: float = 5.0,
    ) -> None:
        self._config = config
        self._ctx = query_context
        self._log = event_log
        self._debounce_s = debounce_s
        self._pending_tasks: dict[str, asyncio.Task[None]] = {}
        self._regen_lock = asyncio.Lock()
        self._config.xlsx_workbooks_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Projection contract
    # ------------------------------------------------------------------
    def apply(self, envelope: dict[str, Any], conn: Any) -> None:
        """Schedule a debounced regeneration for the affected workbook.

        Ignores ``conn`` — this projection has no SQLite backing store.
        xlsx_workbooks is the only projection allowed to write non-DB
        derived state (two files on disk) per [§13.5].
        """
        event_type = envelope.get("type")
        if event_type is None:
            return
        workbook = _TYPE_TO_WORKBOOK.get(event_type)
        if workbook is None:
            return
        self._schedule_regeneration(workbook)

    # ------------------------------------------------------------------
    # Debounce + regeneration
    # ------------------------------------------------------------------
    def _schedule_regeneration(self, workbook: str) -> None:
        """Cancel any pending regenerate for this workbook and schedule a
        fresh one that fires after ``debounce_s`` seconds of silence."""
        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            return

        existing = self._pending_tasks.get(workbook)
        if existing is not None and not existing.done():
            existing.cancel()

        async def _wait_then_run() -> None:
            try:
                await asyncio.sleep(self._debounce_s)
            except asyncio.CancelledError:
                return
            try:
                await self.regenerate_now(workbook)
            except Exception:
                _log.exception(
                    "xlsx_workbooks: regeneration failed for %s", workbook
                )

        self._pending_tasks[workbook] = loop.create_task(_wait_then_run())

    async def regenerate_now(self, workbook: str) -> None:
        """Bypass the debounce: regenerate ``workbook`` immediately and
        emit ``xlsx.regenerated``. Used by tests and by ``adminme
        projection rebuild xlsx_workbooks`` (future prompt)."""
        if workbook not in (OPS_WORKBOOK_NAME, FINANCE_WORKBOOK_NAME):
            raise ValueError(f"unknown workbook: {workbook!r}")

        async with self._regen_lock:
            await self._regenerate(workbook)

    async def _regenerate(self, workbook: str) -> None:
        from adminme.projections.xlsx_workbooks.builders import (
            build_finance_workbook,
            build_ops_workbook,
        )
        from adminme.projections.xlsx_workbooks.lockfile import (
            acquire_workbook_lock,
        )

        start_ms = int(time.time() * 1000)
        path = self._config.xlsx_workbooks_dir / workbook
        lock_path = path.with_suffix(path.suffix + ".lock")
        last_event_id = await self._log.latest_event_id() or ""

        def _under_lock() -> None:
            with acquire_workbook_lock(lock_path, timeout_s=10.0):
                if workbook == OPS_WORKBOOK_NAME:
                    build_ops_workbook(
                        path,
                        self._ctx,
                        tenant_id=self._config.tenant_id,
                        last_event_id=last_event_id,
                    )
                else:
                    build_finance_workbook(
                        path,
                        self._ctx,
                        tenant_id=self._config.tenant_id,
                        last_event_id=last_event_id,
                    )
                # Sidecar must be written INSIDE the same lock as the xlsx
                # write so reverse (07c-β) cannot observe a workbook whose
                # sidecar baseline is from the prior cycle. Reads back from
                # the just-written xlsx (not re-queries projections) so the
                # sidecar is byte-aligned with the workbook on disk.
                self._write_sidecar_for(workbook, path)

        await asyncio.to_thread(_under_lock)

        sheets = (
            _BIDIRECTIONAL_SHEETS[workbook] + _READONLY_SHEETS[workbook]
        )
        duration_ms = int(time.time() * 1000) - start_ms

        # Emit xlsx.regenerated as a SYSTEM event per [§2.2] resolution.
        # This is the ONLY emit in this projection.
        await self._log.append(
            EventEnvelope(
                event_at_ms=int(time.time() * 1000),
                tenant_id=self._config.tenant_id,
                type="xlsx.regenerated",
                schema_version=1,
                occurred_at=EventEnvelope.now_utc_iso(),
                source_adapter="xlsx_workbooks",
                source_account_id="projection",
                owner_scope="shared:household",
                visibility_scope="shared:household",
                sensitivity="normal",
                payload={
                    "workbook_name": workbook,
                    "generated_at": EventEnvelope.now_utc_iso(),
                    "last_event_id_consumed": last_event_id,
                    "sheets_regenerated": sheets,
                    "duration_ms": duration_ms,
                },
            )
        )

    # ------------------------------------------------------------------
    # Sidecar
    # ------------------------------------------------------------------
    def _write_sidecar_for(self, workbook: str, xlsx_path: Path) -> None:
        """Write per-sheet sidecar JSON next to ``xlsx_path``.

        Reads the just-written xlsx (rather than re-querying projection
        databases) so the sidecar is guaranteed to match the workbook
        byte-for-byte. Bidirectional sheets get a row-list payload;
        read-only sheets get a content-hash payload (07c-β WARN signal,
        not a state baseline).
        """
        from openpyxl import load_workbook

        workbooks_dir = self._config.xlsx_workbooks_dir
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=False)
        try:
            for sheet_name in _BIDIRECTIONAL_SHEETS.get(workbook, []):
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    write_sheet_state(workbooks_dir, workbook, sheet_name, [])
                    continue
                headers = [
                    str(h) if h is not None else "" for h in rows[0]
                ]
                row_dicts: list[dict[str, Any]] = []
                for raw in rows[1:]:
                    row_dicts.append(
                        {headers[i]: raw[i] for i in range(len(headers))}
                    )
                write_sheet_state(
                    workbooks_dir, workbook, sheet_name, row_dicts
                )
            for sheet_name in _READONLY_SHEETS.get(workbook, []):
                ws = wb[sheet_name]
                raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
                write_readonly_state(
                    workbooks_dir, workbook, sheet_name, raw_rows
                )
        finally:
            wb.close()


__all__ = [
    "XlsxWorkbooksProjection",
    "XlsxQueryContext",
    "OPS_WORKBOOK_NAME",
    "FINANCE_WORKBOOK_NAME",
]
