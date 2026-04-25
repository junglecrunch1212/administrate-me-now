"""
XlsxReverseDaemon — the reverse xlsx daemon (07c-β).

Per ADMINISTRATEME_BUILD.md §3.11 (lines 993–1080) and SYSTEM_INVARIANTS.md
§§2.2 / 6 / 10 / 13. The daemon is L1-adjacent (NOT a projection): it
watches the workbook on disk via watchdog, debounces a flush window,
acquires the same per-workbook advisory lock the forward daemon uses,
diffs the live workbook against the per-sheet sidecar baseline, and emits
domain events on principal authority. UT-7 (principal_member_id
attribution) is OPEN here and resolves in prompt 08; this prompt stubs
``actor_identity = "xlsx_reverse"`` on every emit and uses the literal
string ``"xlsx_reverse"`` for ``*_by_party_id`` payload fields.

Per-cycle algorithm:
1. Wait ``flush_wait_s`` for writers to flush.
2. Acquire forward lock with ``forward_lock_timeout_s``. On TimeoutError,
   emit ``xlsx.reverse_skipped_during_forward`` and return — skip is the
   cycle terminus, no ``xlsx.reverse_projected`` follows.
3. Open the workbook (data_only=True).
4. For each bidirectional sheet: load rows, read sidecar, diff if sidecar
   exists, dispatch added/updated/deleted to per-sheet emit helpers.
   Cold-start sheets (no sidecar) emit nothing this cycle.
5. For each read-only sheet: hash live rows; WARN if hash drifted.
6. Rewrite all sidecar JSON to current state — even when nothing emitted.
7. Release the lock.
8. Emit ``xlsx.reverse_projected`` with ``events_emitted`` (envelope ids),
   ``sheets_affected`` (sheets that produced ≥1 event or had a non-empty
   diff), and ``duration_ms``.

Undo window: deletes on sheets with ``deletes_use_undo_window=True`` queue
an asyncio.Task that sleeps ``delete_undo_window_s`` then emits the delete
event. A subsequent cycle observing the row return cancels the task. Note:
delete events fired off the undo-window task are NOT included in the
``events_emitted`` of any ``xlsx.reverse_projected`` payload — they emit
after the originating cycle's terminal emit. Acceptable observability gap
documented in BUILD.md §3.11.

Watchdog→asyncio bridge: ``watchdog.observers.Observer`` callbacks run on
watchdog's thread. The daemon hops to its asyncio loop via
``loop.call_soon_threadsafe(self._schedule_cycle, workbook_name)``; the
on-loop ``_schedule_cycle`` debounces (``flush_wait_s`` sleep, cancellable
on subsequent file events) and runs the cycle. Per-workbook serialization
is enforced by an internal ``asyncio.Lock``.
"""

from __future__ import annotations

import asyncio
import logging
import secrets
import time
from pathlib import Path
from typing import Any

import sqlcipher3
from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.diff import DiffResult, diff_sheet
from adminme.daemons.xlsx_sync.sheet_schemas import (
    SheetDescriptor,
    descriptor_for,
)
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import InstanceConfig
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
)
from adminme.projections.xlsx_workbooks.lockfile import acquire_workbook_lock
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext
from adminme.projections.xlsx_workbooks.sidecar import (
    hash_readonly_sheet,
    read_readonly_state,
    read_sheet_state,
    write_readonly_state,
    write_sheet_state,
)

_log = logging.getLogger(__name__)

_BIDIRECTIONAL_SHEETS: dict[str, tuple[str, ...]] = {
    OPS_WORKBOOK_NAME: ("Tasks", "Recurrences", "Commitments"),
    FINANCE_WORKBOOK_NAME: ("Raw Data",),
}
_READONLY_SHEETS: dict[str, tuple[str, ...]] = {
    OPS_WORKBOOK_NAME: ("People", "Metadata"),
    FINANCE_WORKBOOK_NAME: ("Accounts", "Metadata"),
}

_ACTOR = "xlsx_reverse"


def _mint_id(prefix: str) -> str:
    return f"{prefix}{secrets.token_hex(4)}"


class XlsxReverseDaemon:
    """Reverse xlsx daemon. See module docstring for the per-cycle algorithm."""

    def __init__(
        self,
        config: InstanceConfig,
        query_context: XlsxQueryContext,
        *,
        event_log: EventLog,
        flush_wait_s: float = 2.0,
        forward_lock_timeout_s: float = 10.0,
        delete_undo_window_s: float = 5.0,
    ) -> None:
        self._config = config
        self._ctx = query_context
        self._log_ = event_log
        self._flush_wait_s = flush_wait_s
        self._forward_lock_timeout_s = forward_lock_timeout_s
        self._delete_undo_window_s = delete_undo_window_s

        self._observer: Any | None = None
        self._loop: asyncio.AbstractEventLoop | None = None

        self._cycle_locks: dict[str, asyncio.Lock] = {}
        self._debounce_tasks: dict[str, asyncio.Task[None]] = {}

        # (sheet, row_id) -> Task scheduled to emit the delete after the
        # undo window elapses.
        self._pending_deletes: dict[tuple[str, str], asyncio.Task[None]] = {}

        # Test instrumentation: counter of currently-running cycles per
        # workbook (used by tests to assert serialization).
        self._inflight_cycles: dict[str, int] = {
            OPS_WORKBOOK_NAME: 0,
            FINANCE_WORKBOOK_NAME: 0,
        }
        self._max_concurrent_observed = 0

        self._started = False
        self._stopping = False

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------
    async def start(self) -> None:
        """Begin watchdog observation. Idempotent."""
        if self._started:
            return
        self._loop = asyncio.get_running_loop()

        from watchdog.events import FileSystemEventHandler
        from watchdog.observers import Observer

        workbooks_dir = self._config.xlsx_workbooks_dir
        workbooks_dir.mkdir(parents=True, exist_ok=True)

        daemon = self

        class _Handler(FileSystemEventHandler):
            def on_modified(self, event: Any) -> None:
                self._dispatch(event)

            def on_created(self, event: Any) -> None:
                self._dispatch(event)

            def _dispatch(self, event: Any) -> None:
                if event.is_directory:
                    return
                src = Path(str(event.src_path))
                name = src.name
                if name.endswith(".lock"):
                    return
                if name not in (OPS_WORKBOOK_NAME, FINANCE_WORKBOOK_NAME):
                    return
                loop = daemon._loop
                if loop is None:
                    return
                loop.call_soon_threadsafe(daemon._schedule_cycle, name)

        observer = Observer()
        observer.schedule(_Handler(), str(workbooks_dir), recursive=False)
        observer.start()
        self._observer = observer
        self._started = True

    async def stop(self) -> None:
        """Cancel pending cycles and stop the watchdog observer. Idempotent."""
        if self._stopping:
            return
        self._stopping = True
        for task in list(self._debounce_tasks.values()):
            if not task.done():
                task.cancel()
        for task in list(self._pending_deletes.values()):
            if not task.done():
                task.cancel()
        if self._observer is not None:
            self._observer.stop()
            self._observer.join(timeout=2.0)
            self._observer = None
        self._started = False
        self._stopping = False

    # ------------------------------------------------------------------
    # Cycle scheduling
    # ------------------------------------------------------------------
    def _schedule_cycle(self, workbook: str) -> None:
        """On the asyncio loop: cancel any pending debounce for ``workbook``
        and schedule a fresh cycle after the flush wait."""
        existing = self._debounce_tasks.get(workbook)
        if existing is not None and not existing.done():
            existing.cancel()

        async def _wait_then_run() -> None:
            try:
                await asyncio.sleep(self._flush_wait_s)
            except asyncio.CancelledError:
                return
            try:
                await self._run_cycle_internal(workbook, waited=True)
            except Exception:
                _log.exception("xlsx_reverse: cycle failed for %s", workbook)

        loop = self._loop or asyncio.get_event_loop()
        self._debounce_tasks[workbook] = loop.create_task(_wait_then_run())

    async def run_cycle_now(self, workbook: str) -> None:
        """Bypass watchdog + flush wait; drive a full cycle. Used by tests."""
        if workbook not in _BIDIRECTIONAL_SHEETS:
            raise ValueError(f"unknown workbook: {workbook!r}")
        await self._run_cycle_internal(workbook, waited=False)

    # ------------------------------------------------------------------
    # The cycle itself
    # ------------------------------------------------------------------
    async def _run_cycle_internal(self, workbook: str, *, waited: bool) -> None:
        lock = self._cycle_locks.setdefault(workbook, asyncio.Lock())
        async with lock:
            self._inflight_cycles[workbook] += 1
            self._max_concurrent_observed = max(
                self._max_concurrent_observed,
                sum(self._inflight_cycles.values()),
            )
            try:
                await self._cycle_under_serialization(workbook, waited=waited)
            finally:
                self._inflight_cycles[workbook] -= 1

    async def _cycle_under_serialization(
        self, workbook: str, *, waited: bool
    ) -> None:
        del waited  # debounce wait handled by ``_schedule_cycle``
        start_ms = int(time.time() * 1000)
        path = self._config.xlsx_workbooks_dir / workbook
        lock_path = path.with_suffix(path.suffix + ".lock")

        events_emitted: list[str] = []
        sheets_affected: list[str] = []

        # Sentinel value returned by ``_under_lock`` when the lock could
        # not be acquired within the configured timeout.
        SKIP = object()

        def _under_lock() -> Any:
            diffs_for_post: list[tuple[str, DiffResult]] = []
            try:
                cm = acquire_workbook_lock(
                    lock_path, timeout_s=self._forward_lock_timeout_s
                )
            except TimeoutError:
                return SKIP
            try:
                with cm:
                    if not path.exists():
                        return diffs_for_post
                    wb = load_workbook(str(path), data_only=True)
                    try:
                        for sheet_name in _BIDIRECTIONAL_SHEETS.get(workbook, ()):
                            if sheet_name not in wb.sheetnames:
                                continue
                            ws = wb[sheet_name]
                            rows = list(ws.iter_rows(values_only=True))
                            if not rows:
                                current_rows: list[dict[str, Any]] = []
                            else:
                                headers = [
                                    str(h) if h is not None else ""
                                    for h in rows[0]
                                ]
                                current_rows = [
                                    {headers[i]: raw[i] for i in range(len(headers))}
                                    for raw in rows[1:]
                                ]
                            sidecar_rows = read_sheet_state(
                                self._config.xlsx_workbooks_dir,
                                workbook,
                                sheet_name,
                            )
                            descriptor = descriptor_for(workbook, sheet_name)
                            if descriptor is None:
                                continue
                            if sidecar_rows is None:
                                diffs_for_post.append((sheet_name, DiffResult()))
                            else:
                                result = diff_sheet(
                                    current_rows, sidecar_rows, descriptor
                                )
                                diffs_for_post.append((sheet_name, result))
                            write_sheet_state(
                                self._config.xlsx_workbooks_dir,
                                workbook,
                                sheet_name,
                                current_rows,
                            )
                        for sheet_name in _READONLY_SHEETS.get(workbook, ()):
                            if sheet_name not in wb.sheetnames:
                                continue
                            ws = wb[sheet_name]
                            raw_rows = [
                                list(r) for r in ws.iter_rows(values_only=True)
                            ]
                            live_hash = hash_readonly_sheet(raw_rows)
                            prior_hash = read_readonly_state(
                                self._config.xlsx_workbooks_dir,
                                workbook,
                                sheet_name,
                            )
                            if prior_hash is not None and prior_hash != live_hash:
                                _log.warning(
                                    "xlsx_reverse: read-only sheet %s/%s "
                                    "hash drifted (principal edited a "
                                    "read-only sheet); no domain event emitted",
                                    workbook,
                                    sheet_name,
                                )
                            write_readonly_state(
                                self._config.xlsx_workbooks_dir,
                                workbook,
                                sheet_name,
                                raw_rows,
                            )
                    finally:
                        wb.close()
            except TimeoutError:
                return SKIP
            return diffs_for_post

        result_obj = await asyncio.to_thread(_under_lock)
        if result_obj is SKIP:
            await self._emit_skip(workbook)
            return
        diffs: list[tuple[str, DiffResult]] = result_obj

        # Cancel pending undo-window deletes whose row reappeared this cycle.
        for sheet_name, diff_result in diffs:
            descriptor = descriptor_for(workbook, sheet_name)
            if descriptor is None or not descriptor.deletes_use_undo_window:
                continue
            current_ids = {
                str(r.get(descriptor.id_column))
                for r in diff_result.added
                if r.get(descriptor.id_column)
            } | {
                str(r.get(descriptor.id_column))
                for r, _changes in diff_result.updated
                if r.get(descriptor.id_column)
            }
            for rid in current_ids:
                self.cancel_pending_delete(sheet_name, rid)

        # Emit domain events from observed diffs (lock released).
        for sheet_name, diff_result in diffs:
            descriptor = descriptor_for(workbook, sheet_name)
            if descriptor is None:
                continue
            if not (diff_result.added or diff_result.updated or diff_result.deleted):
                continue
            sheet_event_ids = await self._emit_diff(workbook, descriptor, diff_result)
            events_emitted.extend(sheet_event_ids)
            if (
                sheet_event_ids
                or diff_result.added
                or diff_result.updated
                or diff_result.deleted
            ):
                sheets_affected.append(sheet_name)

        # Step 8 (continued): emit xlsx.reverse_projected.
        duration_ms = max(0, int(time.time() * 1000) - start_ms)
        await self._log_.append(
            EventEnvelope(
                event_at_ms=int(time.time() * 1000),
                tenant_id=self._config.tenant_id,
                type="xlsx.reverse_projected",
                schema_version=1,
                occurred_at=EventEnvelope.now_utc_iso(),
                source_adapter="xlsx_reverse",
                source_account_id="daemon",
                owner_scope="shared:household",
                visibility_scope="shared:household",
                sensitivity="normal",
                actor_identity=_ACTOR,
                payload={
                    "workbook_name": workbook,
                    "detected_at": EventEnvelope.now_utc_iso(),
                    "sheets_affected": sheets_affected,
                    "events_emitted": events_emitted,
                    "duration_ms": duration_ms,
                },
            )
        )

    async def _emit_skip(self, workbook: str) -> None:
        await self._log_.append(
            EventEnvelope(
                event_at_ms=int(time.time() * 1000),
                tenant_id=self._config.tenant_id,
                type="xlsx.reverse_skipped_during_forward",
                schema_version=1,
                occurred_at=EventEnvelope.now_utc_iso(),
                source_adapter="xlsx_reverse",
                source_account_id="daemon",
                owner_scope="shared:household",
                visibility_scope="shared:household",
                sensitivity="normal",
                actor_identity=_ACTOR,
                payload={
                    "workbook_name": workbook,
                    "detected_at": EventEnvelope.now_utc_iso(),
                    "skip_reason": "forward_lock_held",
                },
            )
        )

    # ------------------------------------------------------------------
    # Per-sheet emit dispatch
    # ------------------------------------------------------------------
    async def _emit_diff(
        self,
        workbook: str,
        descriptor: SheetDescriptor,
        result: DiffResult,
    ) -> list[str]:
        emitted: list[str] = []
        sheet = descriptor.sheet
        if sheet == "Tasks":
            emitted.extend(await self._emit_tasks(result))
        elif sheet == "Commitments":
            emitted.extend(await self._emit_commitments(result))
        elif sheet == "Recurrences":
            emitted.extend(await self._emit_recurrences(result))
        elif sheet == "Raw Data":
            emitted.extend(await self._emit_raw_data(result))
        else:
            _log.info(
                "[reverse] no emit pathway for sheet %s/%s; "
                "%d adds, %d updates, %d deletes ignored",
                workbook,
                sheet,
                len(result.added),
                len(result.updated),
                len(result.deleted),
            )
        return emitted

    # ------------------------------------------------------------------
    # Tasks pathway
    # ------------------------------------------------------------------
    async def _emit_tasks(self, result: DiffResult) -> list[str]:
        emitted: list[str] = []
        for row in result.added:
            task_id = row.get("task_id") or ""
            if task_id == "" or task_id is None:
                task_id = _mint_id("tsk_")
            payload: dict[str, Any] = {
                "task_id": task_id,
                "title": row.get("title") or "",
            }
            notes = row.get("notes")
            if notes:
                payload["description"] = notes
            assigned = row.get("assigned_member")
            if assigned:
                payload["owner_member_id"] = assigned
            due = row.get("due_date")
            if due:
                payload["due"] = self._stringify_date(due)
            energy = row.get("energy")
            if energy in ("low", "medium", "high"):
                payload["energy"] = energy
            if not payload["title"]:
                _log.info("[reverse] tasks ADD with blank title; skipping")
                continue
            event_id = await self._append(
                event_type="task.created",
                payload=payload,
                sensitivity="normal",
            )
            emitted.append(event_id)

        for row, changes in result.updated:
            task_id = row.get("task_id")
            if not task_id:
                continue
            field_updates = {col: cv for col, (_sv, cv) in changes.items()}
            payload = {
                "task_id": task_id,
                "updated_at": EventEnvelope.now_utc_iso(),
                "updated_by_party_id": None,
                "field_updates": field_updates,
            }
            sensitivity = self._lookup_sensitivity(
                self._ctx.tasks_conn, "tasks", "task_id", task_id
            )
            event_id = await self._append(
                event_type="task.updated",
                payload=payload,
                sensitivity=sensitivity,
            )
            emitted.append(event_id)

        for row in result.deleted:
            task_id = row.get("task_id")
            if not task_id:
                continue
            self._schedule_undo_delete(
                sheet="Tasks",
                row_id=str(task_id),
                emit_factory=lambda tid=str(task_id): self._emit_task_deleted(tid),
            )
        return emitted

    async def _emit_task_deleted(self, task_id: str) -> None:
        sensitivity = self._lookup_sensitivity(
            self._ctx.tasks_conn, "tasks", "task_id", task_id
        )
        await self._append(
            event_type="task.deleted",
            payload={
                "task_id": task_id,
                "deleted_at": EventEnvelope.now_utc_iso(),
                "deleted_by_party_id": _ACTOR,
            },
            sensitivity=sensitivity,
        )

    # ------------------------------------------------------------------
    # Commitments pathway
    # ------------------------------------------------------------------
    async def _emit_commitments(self, result: DiffResult) -> list[str]:
        emitted: list[str] = []
        if result.added:
            _log.info(
                "[reverse] commitments are pipeline-proposed only per [§4.2]; "
                "dropping %d added rows",
                len(result.added),
            )
        for row, changes in result.updated:
            commitment_id = row.get("commitment_id")
            if not commitment_id:
                continue
            field_updates = {col: cv for col, (_sv, cv) in changes.items()}
            sensitivity = self._lookup_sensitivity(
                self._ctx.commitments_conn,
                "commitments",
                "commitment_id",
                commitment_id,
            )
            event_id = await self._append(
                event_type="commitment.edited",
                payload={
                    "commitment_id": commitment_id,
                    "edited_at": EventEnvelope.now_utc_iso(),
                    "edited_by_party_id": _ACTOR,
                    "field_updates": field_updates,
                },
                sensitivity=sensitivity,
            )
            emitted.append(event_id)
        if result.deleted:
            _log.info(
                "[reverse] commitments cancel via API; dropping %d row deletions",
                len(result.deleted),
            )
        return emitted

    # ------------------------------------------------------------------
    # Recurrences pathway
    # ------------------------------------------------------------------
    async def _emit_recurrences(self, result: DiffResult) -> list[str]:
        emitted: list[str] = []
        for row in result.added:
            recurrence_id = row.get("recurrence_id") or ""
            if recurrence_id == "" or recurrence_id is None:
                recurrence_id = _mint_id("rec_")
            cadence = row.get("cadence") or "weekly"
            kind = row.get("title") or "recurrence"
            next_due = row.get("next_due")
            if next_due:
                next_occurrence = self._stringify_date(next_due)
            else:
                next_occurrence = EventEnvelope.now_utc_iso().split("T")[0]
            payload: dict[str, Any] = {
                "recurrence_id": recurrence_id,
                "linked_kind": "household",
                "linked_id": "household",
                "kind": kind,
                "rrule": cadence,
                "next_occurrence": next_occurrence,
            }
            notes = row.get("notes")
            if notes:
                payload["notes"] = notes
            event_id = await self._append(
                event_type="recurrence.added",
                payload=payload,
                sensitivity="normal",
            )
            emitted.append(event_id)

        for row, changes in result.updated:
            recurrence_id = row.get("recurrence_id")
            if not recurrence_id:
                continue
            field_updates = {col: cv for col, (_sv, cv) in changes.items()}
            sensitivity = self._lookup_sensitivity(
                self._ctx.recurrences_conn,
                "recurrences",
                "recurrence_id",
                recurrence_id,
            )
            event_id = await self._append(
                event_type="recurrence.updated",
                payload={
                    "recurrence_id": recurrence_id,
                    "updated_at": EventEnvelope.now_utc_iso(),
                    "field_updates": field_updates,
                },
                sensitivity=sensitivity,
            )
            emitted.append(event_id)

        if result.deleted:
            _log.info(
                "[reverse] recurrences not deletable in v1 per descriptor; "
                "dropping %d row deletions",
                len(result.deleted),
            )
        return emitted

    # ------------------------------------------------------------------
    # Raw Data pathway
    # ------------------------------------------------------------------
    async def _emit_raw_data(self, result: DiffResult) -> list[str]:
        emitted: list[str] = []
        for row in result.added:
            is_manual = bool(row.get("is_manual"))
            if not is_manual:
                _log.warning(
                    "[reverse] non-manual row added via xlsx; rejecting per [§13.4]"
                )
                continue
            flow_id = row.get("txn_id") or ""
            if flow_id == "" or flow_id is None:
                flow_id = _mint_id("flow_")
            amount = row.get("amount")
            try:
                amount_minor = int(round(float(amount) * 100)) if amount is not None else 0
            except (TypeError, ValueError):
                _log.warning("[reverse] manual row amount unparseable; skipping")
                continue
            occurred = row.get("date")
            occurred_at = (
                self._stringify_date(occurred)
                if occurred
                else EventEnvelope.now_utc_iso()
            )
            payload: dict[str, Any] = {
                "flow_id": flow_id,
                "amount_minor": amount_minor,
                "currency": "USD",
                "occurred_at": occurred_at,
                "kind": "paid",
                "added_by_party_id": _ACTOR,
            }
            assigned = row.get("assigned_category")
            if assigned:
                payload["category"] = assigned
            notes = row.get("notes")
            if notes:
                payload["notes"] = notes
            event_id = await self._append(
                event_type="money_flow.manually_added",
                payload=payload,
                sensitivity="normal",
            )
            emitted.append(event_id)

        if result.updated:
            _log.info(
                "[reverse] money_flow.recategorized not registered yet; "
                "dropping %d Raw Data updates (descriptor updates_emit_event=None)",
                len(result.updated),
            )

        for row in result.deleted:
            is_manual = bool(row.get("is_manual"))
            if not is_manual:
                _log.warning(
                    "[reverse] Plaid row deletion via xlsx ignored"
                )
                continue
            flow_id = row.get("txn_id")
            if not flow_id:
                continue
            self._schedule_undo_delete(
                sheet="Raw Data",
                row_id=str(flow_id),
                emit_factory=lambda fid=str(flow_id): self._emit_money_flow_deleted(fid),
            )
        return emitted

    async def _emit_money_flow_deleted(self, flow_id: str) -> None:
        sensitivity = self._lookup_sensitivity(
            self._ctx.money_conn, "money_flows", "flow_id", flow_id
        )
        await self._append(
            event_type="money_flow.manually_deleted",
            payload={
                "flow_id": flow_id,
                "deleted_at": EventEnvelope.now_utc_iso(),
                "deleted_by_party_id": _ACTOR,
            },
            sensitivity=sensitivity,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _schedule_undo_delete(
        self,
        *,
        sheet: str,
        row_id: str,
        emit_factory: Any,
    ) -> None:
        """Queue the delete to fire after ``delete_undo_window_s``. If the
        same (sheet, row_id) is observed back in a subsequent cycle, the
        scheduled task is cancelled."""
        key = (sheet, row_id)
        existing = self._pending_deletes.get(key)
        if existing is not None and not existing.done():
            existing.cancel()

        async def _wait_then_emit() -> None:
            try:
                await asyncio.sleep(self._delete_undo_window_s)
            except asyncio.CancelledError:
                return
            try:
                await emit_factory()
            except Exception:
                _log.exception(
                    "xlsx_reverse: undo-window emit failed for %s/%s",
                    sheet,
                    row_id,
                )
            finally:
                self._pending_deletes.pop(key, None)

        loop = self._loop or asyncio.get_event_loop()
        self._pending_deletes[key] = loop.create_task(_wait_then_emit())

    def cancel_pending_delete(self, sheet: str, row_id: str) -> bool:
        """Cancel a pending undo-window delete if present. Returns True if
        a task was cancelled. Used implicitly by subsequent cycles when a
        diffed row reappears within the undo window."""
        task = self._pending_deletes.pop((sheet, row_id), None)
        if task is None:
            return False
        if not task.done():
            task.cancel()
            return True
        return False

    @staticmethod
    def _stringify_date(value: Any) -> str:
        from datetime import date, datetime as dt

        if isinstance(value, dt):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

    def _lookup_sensitivity(
        self,
        conn: sqlcipher3.Connection,
        table: str,
        id_column: str,
        row_id: str,
    ) -> str:
        try:
            cur = conn.execute(
                f"SELECT sensitivity FROM {table} "
                f"WHERE tenant_id = ? AND {id_column} = ?",
                (self._config.tenant_id, row_id),
            )
            row = cur.fetchone()
        except Exception:
            return "normal"
        if row is None:
            return "normal"
        try:
            value = row["sensitivity"]
        except (KeyError, IndexError, TypeError):
            value = row[0] if isinstance(row, (tuple, list)) else None
        return str(value) if value else "normal"

    async def _append(
        self,
        *,
        event_type: str,
        payload: dict[str, Any],
        sensitivity: str,
    ) -> str:
        envelope = EventEnvelope(
            event_at_ms=int(time.time() * 1000),
            tenant_id=self._config.tenant_id,
            type=event_type,
            schema_version=1,
            occurred_at=EventEnvelope.now_utc_iso(),
            source_adapter="xlsx_reverse",
            source_account_id="daemon",
            owner_scope="shared:household",
            visibility_scope="shared:household",
            sensitivity=sensitivity,
            actor_identity=_ACTOR,
            payload=payload,
        )
        return await self._log_.append(envelope)


__all__ = ["XlsxReverseDaemon"]
