"""
Unit tests for the forward xlsx daemon writing per-sheet sidecar JSON
inside the regeneration lock (07c-α-3).

Per BUILD.md §3.11 lines 1009 / 1015 / 1054 — the reverse daemon (07c-β)
diffs the live workbook against this sidecar to distinguish principal
edits from forward-regenerations. The sidecar must be on disk by the
time ``regenerate_now`` returns and must match the workbook's contents
byte-for-byte (not the projection-DB state — they can drift slightly
across the regenerate boundary).
"""

from __future__ import annotations

import asyncio
import time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext
from adminme.projections.xlsx_workbooks.sidecar import (
    read_readonly_state,
    read_sheet_state,
    sidecar_path,
)

TEST_KEY = b"s" * 32


def _envelope(event_type: str, payload: dict[str, Any]) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, subscriber_id: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        status = await bus.subscriber_status(subscriber_id)
        if status["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"subscriber {subscriber_id} stayed lagged: {status}")


@pytest.fixture
async def rig(tmp_path: Path):
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    # Seed: 1 person, 1 task, 1 recurrence, 1 commitment, 1 account + 2 money flows.
    await log.append(_envelope("party.created", {
        "party_id": "p1", "kind": "person",
        "display_name": "Person One", "sort_name": "Person One",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t1", "title": "first task",
    }))
    await log.append(_envelope("commitment.proposed", {
        "commitment_id": "c1",
        "kind": "reply",
        "owed_by_member_id": "p1",
        "owed_to_party_id": "p1",
        "text_summary": "say hi",
        "confidence": 0.9,
        "strength": "confident",
    }))
    await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r1",
        "linked_kind": "party",
        "linked_id": "p1",
        "kind": "checkin",
        "rrule": "FREQ=WEEKLY",
        "next_occurrence": "2026-05-01T09:00:00Z",
    }))
    await log.append(_envelope("account.added", {
        "account_id": "a1", "display_name": "Checking",
        "organization_party_id": "p1",
        "kind": "bank",
        "status": "active",
        "attributes": {"last4": "1234"},
    }))
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "f1",
        "amount_minor": 1234,
        "currency": "USD",
        "occurred_at": "2026-04-20T10:00:00Z",
        "kind": "paid",
        "linked_account": "a1",
        "source_adapter": "plaid",
    }))
    last = await log.append(_envelope("money_flow.manually_added", {
        "flow_id": "f2",
        "amount_minor": 4321,
        "currency": "USD",
        "occurred_at": "2026-04-21T10:00:00Z",
        "kind": "paid",
        "category": "groceries",
        "notes": "test note",
        "added_by_party_id": "p1",
    }))
    assert last is not None
    await bus.notify(last)
    for name in (
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    projection = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.1)
    try:
        yield {
            "config": config, "log": log, "bus": bus, "runner": runner,
            "ctx": ctx, "projection": projection,
        }
    finally:
        await runner.stop()
        await log.close()


async def test_bidirectional_sheets_get_json_sidecar(rig: dict[str, Any]) -> None:
    """For each bidirectional sheet, sidecar exists with a ``rows`` list."""
    config = rig["config"]
    projection = rig["projection"]
    workbooks_dir = config.xlsx_workbooks_dir

    await projection.regenerate_now(OPS_WORKBOOK_NAME)

    for sheet in ("Tasks", "Recurrences", "Commitments"):
        path = sidecar_path(workbooks_dir, OPS_WORKBOOK_NAME, sheet)
        assert path.exists(), f"sidecar missing for {sheet}"
        rows = read_sheet_state(workbooks_dir, OPS_WORKBOOK_NAME, sheet)
        assert rows is not None
        assert isinstance(rows, list)


async def test_readonly_sheets_get_hash_sidecar(rig: dict[str, Any]) -> None:
    """For each read-only sheet, sidecar has ``content_hash`` only."""
    import json

    config = rig["config"]
    projection = rig["projection"]
    workbooks_dir = config.xlsx_workbooks_dir

    await projection.regenerate_now(OPS_WORKBOOK_NAME)

    for sheet in ("People", "Metadata"):
        path = sidecar_path(workbooks_dir, OPS_WORKBOOK_NAME, sheet)
        assert path.exists(), f"sidecar missing for {sheet}"
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "content_hash" in data
        assert "rows" not in data
        h = data["content_hash"]
        assert isinstance(h, str)
        assert len(h) == 64
        # Confirm the read helper agrees.
        assert read_readonly_state(workbooks_dir, OPS_WORKBOOK_NAME, sheet) == h


async def test_finance_raw_data_sidecar_matches_workbook_rows(
    rig: dict[str, Any],
) -> None:
    """Round-trip canary: sidecar Raw Data rows match the just-written xlsx."""
    config = rig["config"]
    projection = rig["projection"]
    workbooks_dir = config.xlsx_workbooks_dir

    await projection.regenerate_now(FINANCE_WORKBOOK_NAME)

    finance_path = workbooks_dir / FINANCE_WORKBOOK_NAME
    wb = load_workbook(str(finance_path), data_only=True)
    try:
        ws = wb["Raw Data"]
        all_rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        expected = [
            {headers[i]: raw[i] for i in range(len(headers))}
            for raw in all_rows[1:]
        ]
    finally:
        wb.close()

    sidecar_rows = read_sheet_state(workbooks_dir, FINANCE_WORKBOOK_NAME, "Raw Data")
    assert sidecar_rows == expected
    # Sanity: actually have data so the comparison isn't vacuous.
    assert len(sidecar_rows) >= 2


async def test_sidecar_exists_when_regenerate_returns(rig: dict[str, Any]) -> None:
    """The sidecar tree is on disk by the time regenerate_now resolves —
    not pending in some background task."""
    config = rig["config"]
    projection = rig["projection"]
    workbooks_dir = config.xlsx_workbooks_dir

    await projection.regenerate_now(OPS_WORKBOOK_NAME)

    # Every bidirectional + read-only sheet sidecar exists at the moment
    # the await returns. No sleeps, no polling.
    for sheet in ("Tasks", "Recurrences", "Commitments", "People", "Metadata"):
        path = sidecar_path(workbooks_dir, OPS_WORKBOOK_NAME, sheet)
        assert path.exists(), f"sidecar for {sheet} not present after await"
