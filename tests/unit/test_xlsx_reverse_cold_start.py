"""
Unit tests for xlsx reverse cold-start handling (07c-β-2).

When a sheet's sidecar JSON does not exist, the daemon must NOT emit
domain events for that sheet's diff — even if the live workbook
"differs" from an empty baseline. The cycle still rewrites the sidecar
so subsequent cycles can diff cleanly. ``xlsx.reverse_projected`` always
emits at the end.
"""

from __future__ import annotations

import asyncio
import shutil
import time
from pathlib import Path
from typing import Any, AsyncIterator

import pytest
from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.reverse import XlsxReverseDaemon
from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext
from adminme.projections.xlsx_workbooks.sidecar import (
    sidecar_dir,
    sidecar_path,
)

TEST_KEY = b"c" * 32


def _envelope(event_type: str, payload: dict[str, Any]) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"{sid} stayed lagged: {s}")


async def _events_of_type(
    log: EventLog, event_type: str, after: str | None
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    async for e in log.read_since(after):
        if e["type"] == event_type:
            out.append(e)
    return out


@pytest.fixture
async def rig(tmp_path: Path) -> AsyncIterator[dict[str, Any]]:
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    await log.append(_envelope("party.created", {
        "party_id": "p1", "kind": "person",
        "display_name": "P1", "sort_name": "P1",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t1", "title": "first task",
    }))
    last = await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r1",
        "linked_kind": "household",
        "linked_id": "household",
        "kind": "maintenance",
        "rrule": "FREQ=MONTHLY",
        "next_occurrence": "2026-05-01T09:00:00Z",
    }))
    await bus.notify(last)
    for name in (
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    forward = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.05)
    await forward.regenerate_now(OPS_WORKBOOK_NAME)
    reverse = XlsxReverseDaemon(
        config,
        ctx,
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=0.5,
        delete_undo_window_s=0.05,
    )
    cursor = await log.latest_event_id()
    try:
        yield {
            "config": config,
            "log": log,
            "ctx": ctx,
            "forward": forward,
            "reverse": reverse,
            "cursor": cursor,
        }
    finally:
        await reverse.stop()
        await runner.stop()
        await log.close()


async def test_no_sidecar_writes_sidecar_emits_nothing(
    rig: dict[str, Any]
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    state_dir = sidecar_dir(config.xlsx_workbooks_dir)
    assert state_dir.exists()
    shutil.rmtree(state_dir)
    assert not state_dir.exists()

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    assert state_dir.exists()
    # All sheets in the workbook should have a sidecar after the cycle.
    for sheet in ("Tasks", "Recurrences", "Commitments", "People", "Metadata"):
        assert sidecar_path(
            config.xlsx_workbooks_dir, OPS_WORKBOOK_NAME, sheet
        ).exists(), f"sidecar missing after cold-start cycle: {sheet}"

    for ev_type in (
        "task.created",
        "task.updated",
        "task.deleted",
        "recurrence.added",
        "recurrence.updated",
        "commitment.edited",
    ):
        evs = await _events_of_type(log, ev_type, cursor)
        assert evs == [], f"unexpected domain emit on cold-start: {ev_type}"

    projected = await _events_of_type(log, "xlsx.reverse_projected", cursor)
    assert len(projected) == 1
    assert projected[0]["payload"]["events_emitted"] == []


async def test_partial_sidecar_only_diffs_what_can_diff(
    rig: dict[str, Any]
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    # Delete only the Tasks sidecar — Recurrences keeps its baseline.
    sidecar_path(
        config.xlsx_workbooks_dir, OPS_WORKBOOK_NAME, "Tasks"
    ).unlink()

    # Edit a Tasks row (would normally produce task.updated) AND a
    # Recurrences row (will produce recurrence.updated, because its
    # sidecar baseline still exists).
    path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
    wb = load_workbook(str(path))
    tasks_ws = wb["Tasks"]
    # Tasks: column 2 = title; first data row is 2.
    tasks_ws.cell(row=2, column=2, value="renamed task")
    rec_ws = wb["Recurrences"]
    # Recurrences: column 6 = notes
    rec_ws.cell(row=2, column=6, value="updated notes")
    wb.save(str(path))
    wb.close()

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    # Tasks sidecar was missing → cold-start, no task.updated emitted.
    task_updates = await _events_of_type(log, "task.updated", cursor)
    assert task_updates == []
    # Recurrences sidecar present → recurrence.updated emitted.
    rec_updates = await _events_of_type(log, "recurrence.updated", cursor)
    assert len(rec_updates) == 1
    assert rec_updates[0]["payload"]["recurrence_id"] == "r1"
