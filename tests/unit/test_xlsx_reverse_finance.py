"""
Unit tests for the xlsx reverse daemon's Raw Data (finance) pathway
(07c-β-2).

Coverage:
- manual row ADD → ``money_flow.manually_added`` (id-prefix shape canary)
- non-manual row ADD → drop with WARN
- manual row DELETE after the undo window → ``money_flow.manually_deleted``
- non-manual row DELETE → drop with WARN, no event
- assigned_category edit on a Plaid row → INFO drop
  (descriptor.updates_emit_event = None at v1)
- amount edit on a manual row → INFO drop (same reason)
"""

from __future__ import annotations

import asyncio
import logging
import re
import time
from pathlib import Path
from typing import Any, AsyncIterator

import pytest
from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.reverse import XlsxReverseDaemon
from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"f" * 32


def _envelope(event_type: str, payload: dict[str, Any]) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"{sid} stayed lagged: {s}")


async def _events_of_type(
    log: EventLog, event_type: str, after: str | None
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    async for e in log.read_since(after):
        if e["type"] == event_type:
            out.append(e)
    return out


def _open_finance(config: Any) -> tuple[Any, Path]:
    path = config.xlsx_workbooks_dir / FINANCE_WORKBOOK_NAME
    wb = load_workbook(str(path))
    return wb, path


def _save(wb: Any, path: Path) -> None:
    wb.save(str(path))
    wb.close()


def _row_index_by_txn_id(ws: Any, txn_id: str) -> int | None:
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0] == txn_id:
            return i
    return None


@pytest.fixture
async def rig(tmp_path: Path) -> AsyncIterator[dict[str, Any]]:
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    # Seed: 1 person, 1 account, 1 Plaid flow, 1 manual flow.
    await log.append(_envelope("party.created", {
        "party_id": "p1", "kind": "person",
        "display_name": "P1", "sort_name": "P1",
    }))
    await log.append(_envelope("account.added", {
        "account_id": "a1", "display_name": "Checking",
        "organization_party_id": "p1", "kind": "bank",
        "status": "active", "attributes": {"last4": "1234"},
    }))
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "fp1", "amount_minor": 1234, "currency": "USD",
        "occurred_at": "2026-04-20T10:00:00Z",
        "kind": "paid", "linked_account": "a1",
        "source_adapter": "plaid",
    }))
    last = await log.append(_envelope("money_flow.manually_added", {
        "flow_id": "fm1", "amount_minor": 4321, "currency": "USD",
        "occurred_at": "2026-04-21T10:00:00Z",
        "kind": "paid",
        "category": "groceries",
        "notes": "manual flow",
        "added_by_party_id": "p1",
    }))
    await bus.notify(last)
    for name in (
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    forward = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.05)
    await forward.regenerate_now(FINANCE_WORKBOOK_NAME)
    reverse = XlsxReverseDaemon(
        config,
        ctx,
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=0.5,
        delete_undo_window_s=0.05,
    )
    cursor = await log.latest_event_id()
    try:
        yield {
            "config": config,
            "log": log,
            "ctx": ctx,
            "forward": forward,
            "reverse": reverse,
            "cursor": cursor,
        }
    finally:
        await reverse.stop()
        await runner.stop()
        await log.close()


async def test_manual_row_add_emits_money_flow_manually_added(
    rig: dict[str, Any]
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value="")  # blank txn_id → mint
    ws.cell(row=new_row, column=2, value="2026-04-25")
    ws.cell(row=new_row, column=6, value=12.50)
    ws.cell(row=new_row, column=9, value="snacks")
    ws.cell(row=new_row, column=10, value="manual snack run")
    ws.cell(row=new_row, column=11, value=True)
    _save(wb, path)

    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)

    added = await _events_of_type(log, "money_flow.manually_added", cursor)
    assert len(added) == 1
    p = added[0]["payload"]
    assert re.match(r"^flow_[0-9a-f]{8}$", p["flow_id"]), p["flow_id"]
    assert p["amount_minor"] == 1250
    assert p["currency"] == "USD"
    assert p["category"] == "snacks"
    assert p["added_by_party_id"] == "xlsx_reverse"


async def test_plaid_row_add_drops_warn(
    rig: dict[str, Any], caplog: pytest.LogCaptureFixture
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value="plaid_x")
    ws.cell(row=new_row, column=2, value="2026-04-25")
    ws.cell(row=new_row, column=6, value=99.99)
    ws.cell(row=new_row, column=11, value=False)  # is_manual=False
    _save(wb, path)

    caplog.set_level(logging.WARNING, logger="adminme.daemons.xlsx_sync.reverse")
    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)

    added = await _events_of_type(log, "money_flow.manually_added", cursor)
    assert added == []
    assert any(
        "non-manual row added via xlsx" in r.message for r in caplog.records
    )


async def test_manual_row_delete_after_window(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    row_idx = _row_index_by_txn_id(ws, "fm1")
    assert row_idx is not None
    ws.delete_rows(row_idx)
    _save(wb, path)

    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)
    await asyncio.sleep(0.2)

    deleted = await _events_of_type(log, "money_flow.manually_deleted", cursor)
    assert len(deleted) == 1
    assert deleted[0]["payload"]["flow_id"] == "fm1"
    assert deleted[0]["payload"]["deleted_by_party_id"] == "xlsx_reverse"


async def test_plaid_row_delete_drops_warn(
    rig: dict[str, Any], caplog: pytest.LogCaptureFixture
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    row_idx = _row_index_by_txn_id(ws, "fp1")
    assert row_idx is not None
    ws.delete_rows(row_idx)
    _save(wb, path)

    caplog.set_level(logging.WARNING, logger="adminme.daemons.xlsx_sync.reverse")
    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)
    await asyncio.sleep(0.2)

    deleted = await _events_of_type(log, "money_flow.manually_deleted", cursor)
    assert deleted == []
    assert any(
        "Plaid row deletion via xlsx ignored" in r.message
        for r in caplog.records
    )


async def test_assigned_category_edit_drops_info(
    rig: dict[str, Any], caplog: pytest.LogCaptureFixture
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    row_idx = _row_index_by_txn_id(ws, "fp1")
    assert row_idx is not None
    # column 9 = assigned_category
    ws.cell(row=row_idx, column=9, value="reclassified")
    _save(wb, path)

    caplog.set_level(logging.INFO, logger="adminme.daemons.xlsx_sync.reverse")
    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)

    # No money_flow.* domain events landed.
    for ev_type in (
        "money_flow.manually_added",
        "money_flow.manually_deleted",
    ):
        evs = await _events_of_type(log, ev_type, cursor)
        assert evs == []
    assert any(
        "money_flow.recategorized not registered" in r.message
        for r in caplog.records
    )


async def test_amount_edit_on_manual_row_drops_info(
    rig: dict[str, Any], caplog: pytest.LogCaptureFixture
) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_finance(config)
    ws = wb["Raw Data"]
    row_idx = _row_index_by_txn_id(ws, "fm1")
    assert row_idx is not None
    # column 6 = amount
    ws.cell(row=row_idx, column=6, value=99.99)
    _save(wb, path)

    caplog.set_level(logging.INFO, logger="adminme.daemons.xlsx_sync.reverse")
    await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)

    for ev_type in (
        "money_flow.manually_added",
        "money_flow.manually_deleted",
    ):
        evs = await _events_of_type(log, ev_type, cursor)
        assert evs == []
    assert any(
        "money_flow.recategorized not registered" in r.message
        for r in caplog.records
    )
