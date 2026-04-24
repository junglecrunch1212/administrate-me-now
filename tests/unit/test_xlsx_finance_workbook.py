"""
Unit tests for the xlsx_workbooks forward daemon — adminme-finance.xlsx.

Per phase 07b prompt Commit 3. Raw Data [bidirectional-shape] + Accounts
[read-only] + Metadata [read-only]. Derived / Plaid-authoritative cells
are locked per the protection matrix in the builder.
"""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks.builders import build_finance_workbook
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"f" * 32


def _envelope(
    event_type: str,
    payload: dict[str, Any],
    *,
    tenant_id: str = "tenant-a",
    owner_scope: str = "shared:household",
) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id=tenant_id,
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope=owner_scope,
        visibility_scope=owner_scope,
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, subscriber_id: str, timeout: float = 10.0) -> None:
    import asyncio

    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        status = await bus.subscriber_status(subscriber_id)
        if status["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"subscriber {subscriber_id} stayed lagged: {status}")


@pytest.fixture
async def populated(tmp_path: Path):
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    # One organization (bank) party
    await log.append(_envelope("party.created", {
        "party_id": "bank1",
        "kind": "organization",
        "display_name": "Bank One",
        "sort_name": "Bank One",
    }))
    # Three accounts
    for i, (kind, name, last4) in enumerate([
        ("bank", "Checking", "1111"),
        ("credit_card", "Card1", "2222"),
        ("brokerage", "Brokerage", "3333"),
    ]):
        await log.append(_envelope("account.added", {
            "account_id": f"a{i}",
            "display_name": name,
            "organization_party_id": "bank1",
            "kind": kind,
            "status": "active",
            "attributes": {"last4": last4},
        }))
    # Two Plaid-recorded flows + one manual + one manual-then-deleted.
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "f1",
        "amount_minor": 1250,
        "currency": "USD",
        "occurred_at": "2026-04-20T10:00:00Z",
        "kind": "paid",
        "category": "groceries",
        "linked_account": "a0",
        "source_adapter": "plaid",
    }))
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "f2",
        "amount_minor": 5500,
        "currency": "USD",
        "occurred_at": "2026-04-21T14:00:00Z",
        "kind": "paid",
        "category": "dining",
        "linked_account": "a1",
        "source_adapter": "plaid",
    }))
    await log.append(_envelope("money_flow.manually_added", {
        "flow_id": "f3",
        "amount_minor": 9999,
        "currency": "USD",
        "occurred_at": "2026-04-22T12:00:00Z",
        "kind": "paid",
        "category": "home_improvement",
        "added_by_party_id": "bank1",
    }))
    await log.append(_envelope("money_flow.manually_added", {
        "flow_id": "f4",
        "amount_minor": 100,
        "currency": "USD",
        "occurred_at": "2026-04-23T09:00:00Z",
        "kind": "paid",
        "added_by_party_id": "bank1",
    }))
    await log.append(_envelope("money_flow.manually_deleted", {
        "flow_id": "f4",
        "deleted_at": "2026-04-23T10:00:00Z",
        "deleted_by_party_id": "bank1",
    }))

    last_eid = await log.latest_event_id()
    assert last_eid is not None
    await bus.notify(last_eid)
    for name in [
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ]:
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    try:
        yield {
            "config": config, "log": log, "bus": bus, "runner": runner,
            "ctx": ctx, "last_event_id": last_eid,
        }
    finally:
        await runner.stop()
        await log.close()


def _column_values(ws: Worksheet, header: str) -> list[Any]:
    headers = [c.value for c in ws[1]]
    col_idx = headers.index(header) + 1
    return [ws.cell(row=i, column=col_idx).value for i in range(2, ws.max_row + 1)]


async def test_raw_data_sheet_shows_plaid_and_manual(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    raw = wb["Raw Data"]
    ids = _column_values(raw, "txn_id")
    # f1 (plaid) + f2 (plaid) + f3 (manual)  — f4 deleted
    assert set(ids) == {"f1", "f2", "f3"}
    is_manual = _column_values(raw, "is_manual")
    assert sum(1 for v in is_manual if v) == 1
    assert sum(1 for v in is_manual if not v) == 2


async def test_raw_data_excludes_soft_deleted(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    raw = wb["Raw Data"]
    assert "f4" not in _column_values(raw, "txn_id")


async def test_raw_data_plaid_cells_locked(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    raw = wb["Raw Data"]
    headers = [c.value for c in raw[1]]
    # Find a plaid row (is_manual=False).
    is_manual_col = headers.index("is_manual") + 1
    plaid_row_idx = None
    for r in range(2, raw.max_row + 1):
        if not raw.cell(row=r, column=is_manual_col).value:
            plaid_row_idx = r
            break
    assert plaid_row_idx is not None
    for col_name in ("date", "account_last4", "amount", "plaid_category", "txn_id"):
        col_idx = headers.index(col_name) + 1
        assert raw.cell(row=plaid_row_idx, column=col_idx).protection.locked is True, col_name
    assigned_col = headers.index("assigned_category") + 1
    assert raw.cell(row=plaid_row_idx, column=assigned_col).protection.locked is False


async def test_raw_data_manual_cells_unlocked(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    raw = wb["Raw Data"]
    headers = [c.value for c in raw[1]]
    is_manual_col = headers.index("is_manual") + 1
    manual_row_idx = None
    for r in range(2, raw.max_row + 1):
        if raw.cell(row=r, column=is_manual_col).value:
            manual_row_idx = r
            break
    assert manual_row_idx is not None
    # txn_id still locked (always derived) but amount/date etc unlocked.
    txn_id_col = headers.index("txn_id") + 1
    assert raw.cell(row=manual_row_idx, column=txn_id_col).protection.locked is True
    for col_name in ("date", "amount", "assigned_category"):
        col_idx = headers.index(col_name) + 1
        assert raw.cell(row=manual_row_idx, column=col_idx).protection.locked is False, col_name


async def test_accounts_sheet_readonly(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    accounts = wb["Accounts"]
    assert accounts.protection.sheet is True
    for row in accounts.iter_rows(min_row=2):
        for cell in row:
            assert cell.protection.locked is True


async def test_accounts_sheet_rows(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    accounts = wb["Accounts"]
    ids = _column_values(accounts, "account_id")
    assert set(ids) == {"a0", "a1", "a2"}
    types = _column_values(accounts, "account_type")
    assert set(types) == {"bank", "credit_card", "brokerage"}
    last4s = _column_values(accounts, "last4")
    assert set(last4s) == {"1111", "2222", "3333"}


async def test_finance_metadata_populated(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-finance.xlsx"
    last_eid = populated["last_event_id"]
    build_finance_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id=last_eid)
    wb = load_workbook(str(path))
    meta = wb["Metadata"]
    pairs = {
        meta.cell(row=i, column=1).value: meta.cell(row=i, column=2).value
        for i in range(2, meta.max_row + 1)
    }
    assert pairs["workbook_name"] == "adminme-finance.xlsx"
    assert pairs["tenant_id"] == config.tenant_id
    assert pairs["last_event_id_consumed"] == last_eid


def _data_rows(ws: Worksheet) -> list[tuple]:
    return [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]


async def test_finance_idempotent_semantically(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path_a = config.xlsx_workbooks_dir / "adminme-finance-a.xlsx"
    path_b = config.xlsx_workbooks_dir / "adminme-finance-b.xlsx"
    build_finance_workbook(path_a, ctx, tenant_id=config.tenant_id, last_event_id="e1")
    build_finance_workbook(path_b, ctx, tenant_id=config.tenant_id, last_event_id="e1")
    wb_a = load_workbook(str(path_a))
    wb_b = load_workbook(str(path_b))
    for name in ("Raw Data", "Accounts"):
        assert _data_rows(wb_a[name]) == _data_rows(wb_b[name]), name
