"""
Unit tests for the xlsx_workbooks forward daemon — adminme-ops.xlsx.

Per phase 07b prompt: sheet builders (Tasks, Recurrences, Commitments,
People, Metadata) produce the expected rows, freeze the header, apply
derived-cell protection, atomic write, lock contention times out, and
regenerate-twice is semantically idempotent.

Tests bypass the 5s debounce by calling ``build_ops_workbook`` directly
(for sheet content assertions) or ``regenerate_now`` (for emit / flow
assertions, which live in test_xlsx_regenerated_emit.py). See CF-5 / CF-7
in the prompt — direct-handler pattern for failure tests.
"""

from __future__ import annotations

import threading
import time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.xlsx_workbooks.builders import build_ops_workbook
from adminme.projections.xlsx_workbooks.lockfile import acquire_workbook_lock
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"x" * 32


def _envelope(
    event_type: str,
    payload: dict[str, Any],
    *,
    tenant_id: str = "tenant-a",
    owner_scope: str = "shared:household",
) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id=tenant_id,
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope=owner_scope,
        visibility_scope=owner_scope,
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, subscriber_id: str, timeout: float = 5.0) -> None:
    import asyncio

    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        status = await bus.subscriber_status(subscriber_id)
        if status["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"subscriber {subscriber_id} stayed lagged: {status}")


@pytest.fixture
async def populated(tmp_path: Path):
    """Run the 7 projections we read from and populate a deterministic
    fixture: 5 parties (2 persons + 2 orgs + 1 household), 3 tasks, 2
    commitments, 2 recurrences. Returns config + connections."""
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    # Parties
    for i in range(2):
        await log.append(_envelope("party.created", {
            "party_id": f"p{i:03d}",
            "kind": "person",
            "display_name": f"Person{i}",
            "sort_name": f"Person{i:03d}",
        }))
    for i in range(2):
        await log.append(_envelope("party.created", {
            "party_id": f"org{i:03d}",
            "kind": "organization",
            "display_name": f"Org{i}",
            "sort_name": f"Org{i:03d}",
        }))
    await log.append(_envelope("party.created", {
        "party_id": "hh",
        "kind": "household",
        "display_name": "HH",
        "sort_name": "HH",
    }))

    # Tasks
    await log.append(_envelope("task.created", {
        "task_id": "t1", "title": "Mow lawn", "owner_member_id": "p000",
        "due": "2026-04-26", "energy": "medium",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t2", "title": "Renew passport", "owner_member_id": "p001",
        "due": "2026-05-10", "energy": "low",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t3", "title": "Schedule dentist", "owner_member_id": "p000",
        "energy": "low",
    }))

    # Commitments
    await log.append(_envelope("commitment.proposed", {
        "commitment_id": "c1",
        "kind": "reply",
        "owed_by_member_id": "p000",
        "owed_to_party_id": "org000",
        "text_summary": "reply to Org0 email",
        "confidence": 0.9,
        "strength": "confident",
    }))
    await log.append(_envelope("commitment.proposed", {
        "commitment_id": "c2",
        "kind": "task",
        "owed_by_member_id": "p001",
        "owed_to_party_id": "p000",
        "text_summary": "pick up prescription",
        "confidence": 0.5,
        "strength": "weak",
    }))

    # Recurrences
    await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r1",
        "linked_kind": "household",
        "linked_id": "hh",
        "kind": "maintenance",
        "rrule": "FREQ=MONTHLY;BYMONTHDAY=1",
        "next_occurrence": "2026-05-01T09:00:00Z",
    }))
    await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r2",
        "linked_kind": "party",
        "linked_id": "p000",
        "kind": "review",
        "rrule": "FREQ=WEEKLY;BYDAY=MO",
        "next_occurrence": "2026-04-27T08:00:00Z",
    }))

    last_eid = await log.latest_event_id()
    assert last_eid is not None
    await bus.notify(last_eid)
    for name in [
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ]:
        await _wait_idle(bus, f"projection:{name}", timeout=10.0)

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )

    try:
        yield {
            "config": config,
            "log": log,
            "bus": bus,
            "runner": runner,
            "ctx": ctx,
            "last_event_id": last_eid or "",
        }
    finally:
        await runner.stop()
        await log.close()


def _data_rows(ws: Worksheet) -> list[tuple]:
    """Return non-header data rows as tuples of cell values."""
    return [
        tuple(cell.value for cell in row)
        for row in ws.iter_rows(min_row=2)
    ]


def _column_values(ws: Worksheet, header: str) -> list[Any]:
    """Return the column of values under ``header`` (data only)."""
    headers = [c.value for c in ws[1]]
    col_idx = headers.index(header) + 1
    return [ws.cell(row=i, column=col_idx).value for i in range(2, ws.max_row + 1)]


async def test_workbook_creates_file(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    assert path.exists()
    assert path.stat().st_size > 0
    wb = load_workbook(str(path))
    assert set(wb.sheetnames) == {
        "Tasks", "Recurrences", "Commitments", "People", "Metadata"
    }


async def test_tasks_sheet_rows(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    tasks = wb["Tasks"]
    assert tasks.max_row >= 4  # header + 3 tasks
    task_ids = _column_values(tasks, "task_id")
    assert set(task_ids) == {"t1", "t2", "t3"}
    titles = _column_values(tasks, "title")
    assert "Mow lawn" in titles
    statuses = _column_values(tasks, "status")
    assert all(s == "inbox" for s in statuses)


async def test_tasks_sheet_headers_frozen(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    assert wb["Tasks"].freeze_panes == "A2"
    assert wb["Recurrences"].freeze_panes == "A2"
    assert wb["Commitments"].freeze_panes == "A2"


async def test_tasks_derived_cells_locked(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    tasks = wb["Tasks"]
    headers = [c.value for c in tasks[1]]
    task_id_col = headers.index("task_id") + 1
    title_col = headers.index("title") + 1
    created_at_col = headers.index("created_at") + 1
    # Row 2 is the first data row.
    assert tasks.cell(row=2, column=task_id_col).protection.locked is True
    assert tasks.cell(row=2, column=created_at_col).protection.locked is True
    # non-derived cells unlocked
    assert tasks.cell(row=2, column=title_col).protection.locked is False


async def test_recurrences_sheet_rows(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    rec = wb["Recurrences"]
    rec_ids = _column_values(rec, "recurrence_id")
    assert set(rec_ids) == {"r1", "r2"}
    cadences = _column_values(rec, "cadence")
    assert any("FREQ=MONTHLY" in (c or "") for c in cadences)


async def test_commitments_sheet_contains_all(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    commit = wb["Commitments"]
    ids = _column_values(commit, "commitment_id")
    assert set(ids) == {"c1", "c2"}
    # strength derived from confidence buckets
    strengths = _column_values(commit, "strength")
    assert "confident" in strengths
    assert "weak" in strengths


async def test_commitments_derived_cells_locked(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    ws = wb["Commitments"]
    headers = [c.value for c in ws[1]]
    for col_name in ("confidence", "strength", "source_summary"):
        col_idx = headers.index(col_name) + 1
        assert ws.cell(row=2, column=col_idx).protection.locked is True, col_name
    # text_summary is editable
    ts_col = headers.index("text_summary") + 1
    assert ws.cell(row=2, column=ts_col).protection.locked is False


async def test_people_sheet_readonly(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="ev_test")
    wb = load_workbook(str(path))
    people = wb["People"]
    assert people.protection.sheet is True
    # excludes household (kind='household')
    party_ids = _column_values(people, "party_id")
    assert "hh" not in party_ids
    # Persons + orgs included
    assert "p000" in party_ids
    assert "org000" in party_ids
    # Every data cell is locked
    for row in people.iter_rows(min_row=2):
        for cell in row:
            assert cell.protection.locked is True


async def test_metadata_sheet_populated(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    last_eid = populated["last_event_id"]
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id=last_eid)
    wb = load_workbook(str(path))
    meta = wb["Metadata"]
    pairs = {
        meta.cell(row=i, column=1).value: meta.cell(row=i, column=2).value
        for i in range(2, meta.max_row + 1)
    }
    assert pairs["tenant_id"] == config.tenant_id
    assert pairs["workbook_name"] == "adminme-ops.xlsx"
    assert pairs["last_event_id_consumed"] == last_eid
    assert "generated_at" in pairs
    assert pairs["generated_at"]  # non-empty
    assert meta.protection.sheet is True


def _sheet_rows_equal(ws1: Worksheet, ws2: Worksheet) -> bool:
    return _data_rows(ws1) == _data_rows(ws2)


async def test_regenerate_idempotent_semantically(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path_a = config.xlsx_workbooks_dir / "adminme-ops-a.xlsx"
    path_b = config.xlsx_workbooks_dir / "adminme-ops-b.xlsx"
    build_ops_workbook(path_a, ctx, tenant_id=config.tenant_id, last_event_id="e1")
    build_ops_workbook(path_b, ctx, tenant_id=config.tenant_id, last_event_id="e1")
    wb_a = load_workbook(str(path_a))
    wb_b = load_workbook(str(path_b))
    for name in ("Tasks", "Recurrences", "Commitments", "People"):
        assert _sheet_rows_equal(wb_a[name], wb_b[name]), f"mismatch in {name}"
    # Metadata: generated_at allowed to differ, but last_event_id_consumed
    # must match.
    meta_a = dict(zip(
        [c.value for c in wb_a["Metadata"]["A"][1:]],
        [c.value for c in wb_a["Metadata"]["B"][1:]],
    ))
    meta_b = dict(zip(
        [c.value for c in wb_b["Metadata"]["A"][1:]],
        [c.value for c in wb_b["Metadata"]["B"][1:]],
    ))
    assert meta_a["last_event_id_consumed"] == meta_b["last_event_id_consumed"]
    assert meta_a["tenant_id"] == meta_b["tenant_id"]


async def test_atomic_write_tmp_file_gone_on_success(populated: dict[str, Any]) -> None:
    config = populated["config"]
    ctx = populated["ctx"]
    path = config.xlsx_workbooks_dir / "adminme-ops.xlsx"
    build_ops_workbook(path, ctx, tenant_id=config.tenant_id, last_event_id="e1")
    leftovers = list(config.xlsx_workbooks_dir.glob("adminme-ops.xlsx.tmp.*"))
    assert leftovers == []


async def test_lock_held_by_another_thread_times_out(populated: dict[str, Any]) -> None:
    config = populated["config"]
    lock_path = config.xlsx_workbooks_dir / "adminme-ops.xlsx.lock"
    config.xlsx_workbooks_dir.mkdir(parents=True, exist_ok=True)
    barrier = threading.Event()
    release = threading.Event()

    def holder() -> None:
        with acquire_workbook_lock(lock_path, timeout_s=5.0):
            barrier.set()
            release.wait()

    t = threading.Thread(target=holder, daemon=True)
    t.start()
    try:
        assert barrier.wait(timeout=3.0), "holder did not acquire lock"
        with pytest.raises(TimeoutError):
            with acquire_workbook_lock(lock_path, timeout_s=0.5):
                pytest.fail("should not have acquired lock")
    finally:
        release.set()
        t.join(timeout=3.0)


async def test_multi_tenant_isolation(tmp_path: Path) -> None:
    """Two tenants populated in separate instance dirs produce separate
    workbook files with separate content, enforced by the tenant filter in
    each sheet builder."""
    instance_a = tmp_path / "a"
    instance_a.mkdir()
    (instance_a / "config").mkdir()
    (instance_a / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    instance_b = tmp_path / "b"
    instance_b.mkdir()
    (instance_b / "config").mkdir()
    (instance_b / "config" / "instance.yaml").write_text("tenant_id: tenant-b\n")

    config_a = load_instance_config(instance_a)
    config_b = load_instance_config(instance_b)

    log_a = EventLog(config_a, TEST_KEY)
    log_b = EventLog(config_b, TEST_KEY)
    bus_a = EventBus(log_a, config_a.bus_checkpoint_path)
    bus_b = EventBus(log_b, config_b.bus_checkpoint_path)
    runner_a = ProjectionRunner(bus_a, log_a, config_a, encryption_key=TEST_KEY)
    runner_b = ProjectionRunner(bus_b, log_b, config_b, encryption_key=TEST_KEY)
    for r in (runner_a, runner_b):
        r.register(PartiesProjection())
        r.register(TasksProjection())
        r.register(CommitmentsProjection())
        r.register(RecurrencesProjection())
        r.register(CalendarsProjection())
        r.register(PlacesAssetsAccountsProjection())
        r.register(MoneyProjection())
    await runner_a.start()
    await runner_b.start()

    try:
        eid_a = await log_a.append(_envelope("task.created", {
            "task_id": "ta", "title": "tenant-a only",
        }, tenant_id="tenant-a"))
        eid_b = await log_b.append(_envelope("task.created", {
            "task_id": "tb", "title": "tenant-b only",
        }, tenant_id="tenant-b"))
        await bus_a.notify(eid_a)
        await bus_b.notify(eid_b)

        await _wait_idle(bus_a, "projection:tasks")
        await _wait_idle(bus_b, "projection:tasks")

        ctx_a = XlsxQueryContext(
            parties_conn=runner_a.connection("parties"),
            tasks_conn=runner_a.connection("tasks"),
            commitments_conn=runner_a.connection("commitments"),
            recurrences_conn=runner_a.connection("recurrences"),
            calendars_conn=runner_a.connection("calendars"),
            places_assets_accounts_conn=runner_a.connection("places_assets_accounts"),
            money_conn=runner_a.connection("money"),
        )
        path_a = config_a.xlsx_workbooks_dir / "adminme-ops.xlsx"
        build_ops_workbook(path_a, ctx_a, tenant_id="tenant-a", last_event_id="ea")
        wb_a = load_workbook(str(path_a))
        ids_a = _column_values(wb_a["Tasks"], "task_id")
        assert ids_a == ["ta"]
        assert "tb" not in ids_a
    finally:
        await runner_a.stop()
        await runner_b.stop()
        await log_a.close()
        await log_b.close()
