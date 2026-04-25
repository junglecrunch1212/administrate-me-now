"""
Unit tests for XlsxReverseDaemon — Tasks pathway + cycle plumbing (07c-β).

Coverage:
- no-op cycle emits only ``xlsx.reverse_projected`` with empty
  ``events_emitted``
- new task row → ``task.created``, including blank-id minting and the
  ``^tsk_[0-9a-f]{8}$`` shape
- title edit → ``task.updated`` with field_updates
- task delete after the undo window elapses → ``task.deleted``
- task delete cancelled within the undo window → no ``task.deleted``
- derived-column edit (``created_at``) is silently dropped
- id-column edit on a Tasks row surfaces as delete-then-add
- read-only sheet edit logs WARN, no domain event, ``xlsx.reverse_projected``
  still emits
- every cycle emits ``xlsx.reverse_projected`` with non-negative
  ``duration_ms``
- sensitivity preservation on update: a sensitive task stays sensitive on
  emitted ``task.updated``
"""

from __future__ import annotations

import asyncio
import re
import time
from pathlib import Path
from typing import Any, AsyncIterator

import pytest
from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.reverse import XlsxReverseDaemon
from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"r" * 32


def _envelope(
    event_type: str,
    payload: dict[str, Any],
    *,
    sensitivity: str = "normal",
) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity=sensitivity,
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 10.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"{sid} stayed lagged: {s}")


async def _read_events(log: EventLog, after: str | None) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    async for e in log.read_since(after):
        out.append(e)
    return out


async def _events_of_type(
    log: EventLog, event_type: str, after: str | None
) -> list[dict[str, Any]]:
    return [e for e in await _read_events(log, after) if e["type"] == event_type]


@pytest.fixture
async def rig(tmp_path: Path) -> AsyncIterator[dict[str, Any]]:
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()

    # Seed: 1 person + 1 task (sensitivity=normal) + 1 task (sensitivity=sensitive)
    await log.append(_envelope("party.created", {
        "party_id": "p1", "kind": "person",
        "display_name": "P1", "sort_name": "P1",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t1", "title": "first task",
    }))
    last = await log.append(
        _envelope(
            "task.created",
            {"task_id": "t_sensitive", "title": "secret task"},
            sensitivity="sensitive",
        )
    )
    await bus.notify(last)
    for name in (
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    forward = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.05)
    await forward.regenerate_now(OPS_WORKBOOK_NAME)
    await forward.regenerate_now(FINANCE_WORKBOOK_NAME)

    reverse = XlsxReverseDaemon(
        config,
        ctx,
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=0.5,
        delete_undo_window_s=0.05,
    )

    cursor = await log.latest_event_id()

    try:
        yield {
            "config": config,
            "log": log,
            "bus": bus,
            "runner": runner,
            "ctx": ctx,
            "forward": forward,
            "reverse": reverse,
            "cursor": cursor,
        }
    finally:
        await reverse.stop()
        await runner.stop()
        await log.close()


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------

def _open_ops(config: Any) -> tuple[Any, Any]:
    path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
    wb = load_workbook(str(path))
    return wb, path


def _save(wb: Any, path: Path) -> None:
    wb.save(str(path))
    wb.close()


def _tasks_sheet(wb: Any) -> Any:
    return wb["Tasks"]


def _row_index_by_task_id(ws: Any, task_id: str) -> int | None:
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0] == task_id:
            return i
    return None


# ----------------------------------------------------------------------
# Tests
# ----------------------------------------------------------------------

async def test_no_edits_no_events(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    cursor = rig["cursor"]

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    new = await _read_events(log, cursor)
    assert [e["type"] for e in new] == ["xlsx.reverse_projected"]
    payload = new[0]["payload"]
    assert payload["events_emitted"] == []
    assert payload["sheets_affected"] == []


async def test_new_task_emits_task_created(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value="t99")
    ws.cell(row=new_row, column=2, value="brand new task")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    created = await _events_of_type(log, "task.created", cursor)
    assert len(created) == 1
    assert created[0]["payload"]["task_id"] == "t99"
    assert created[0]["payload"]["title"] == "brand new task"


async def test_blank_id_gets_generated(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    new_row = ws.max_row + 1
    # Leave task_id blank.
    ws.cell(row=new_row, column=2, value="needs an id")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    created = await _events_of_type(log, "task.created", cursor)
    assert len(created) == 1
    minted = created[0]["payload"]["task_id"]
    assert re.match(r"^tsk_[0-9a-f]{8}$", minted), minted


async def test_edit_title_emits_task_updated(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    row_idx = _row_index_by_task_id(ws, "t1")
    assert row_idx is not None
    ws.cell(row=row_idx, column=2, value="renamed task")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    updated = await _events_of_type(log, "task.updated", cursor)
    assert len(updated) == 1
    p = updated[0]["payload"]
    assert p["task_id"] == "t1"
    assert p["field_updates"] == {"title": "renamed task"}


async def test_delete_after_undo_window(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    row_idx = _row_index_by_task_id(ws, "t1")
    assert row_idx is not None
    ws.delete_rows(row_idx)
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
    # Wait for the undo window plus a small safety margin to pass.
    await asyncio.sleep(0.2)

    deleted = await _events_of_type(log, "task.deleted", cursor)
    assert len(deleted) == 1
    assert deleted[0]["payload"]["task_id"] == "t1"


async def test_delete_cancelled_within_window(rig: dict[str, Any]) -> None:
    config = rig["config"]
    log = rig["log"]
    cursor = rig["cursor"]
    # Use a longer undo window so we can race a second cycle in.
    reverse = XlsxReverseDaemon(
        config,
        rig["ctx"],
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=0.5,
        delete_undo_window_s=2.0,
    )
    try:
        wb, path = _open_ops(config)
        ws = _tasks_sheet(wb)
        # Capture the original t1 row's contents before deleting.
        row_idx = _row_index_by_task_id(ws, "t1")
        assert row_idx is not None
        original = [ws.cell(row=row_idx, column=c).value for c in range(1, 14)]
        ws.delete_rows(row_idx)
        _save(wb, path)

        await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
        # Restore the row before the undo window elapses.
        wb, path = _open_ops(config)
        ws = _tasks_sheet(wb)
        new_row = ws.max_row + 1
        for c, v in enumerate(original, start=1):
            ws.cell(row=new_row, column=c, value=v)
        _save(wb, path)

        await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
        # Wait long enough for the undo window to elapse.
        await asyncio.sleep(2.5)

        deleted = await _events_of_type(log, "task.deleted", cursor)
        assert deleted == []
    finally:
        await reverse.stop()


async def test_derived_column_dropped(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    row_idx = _row_index_by_task_id(ws, "t1")
    assert row_idx is not None
    # created_at is column 12.
    ws.cell(row=row_idx, column=12, value="2099-01-01T00:00:00Z")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    updated = await _events_of_type(log, "task.updated", cursor)
    assert updated == []
    # Subsequent cycle should emit nothing further (sidecar is rewritten).
    cursor2 = await log.latest_event_id()
    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
    new = await _events_of_type(log, "task.updated", cursor2)
    assert new == []


async def test_id_column_change_as_delete_plus_add(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    row_idx = _row_index_by_task_id(ws, "t1")
    assert row_idx is not None
    ws.cell(row=row_idx, column=1, value="t99")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
    await asyncio.sleep(0.2)

    created = await _events_of_type(log, "task.created", cursor)
    deleted = await _events_of_type(log, "task.deleted", cursor)
    assert len(created) == 1
    assert created[0]["payload"]["task_id"] == "t99"
    assert len(deleted) == 1
    assert deleted[0]["payload"]["task_id"] == "t1"


async def test_readonly_sheet_warn(
    rig: dict[str, Any], caplog: pytest.LogCaptureFixture
) -> None:
    import logging

    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    people_ws = wb["People"]
    # Add a junk cell in the People sheet to drift its hash.
    people_ws.cell(row=people_ws.max_row + 1, column=1, value="X")
    _save(wb, path)

    caplog.set_level(logging.WARNING, logger="adminme.daemons.xlsx_sync.reverse")
    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    new = await _read_events(log, cursor)
    types = [e["type"] for e in new]
    # No domain event landed; xlsx.reverse_projected still emitted.
    assert types == ["xlsx.reverse_projected"]
    assert any("hash drifted" in r.message for r in caplog.records)


async def test_reverse_projected_at_cycle_end(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    cursor = rig["cursor"]

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
    rp = await _events_of_type(log, "xlsx.reverse_projected", cursor)
    assert len(rp) == 1
    assert rp[0]["payload"]["duration_ms"] >= 0


async def test_sensitivity_preserved_on_update(rig: dict[str, Any]) -> None:
    reverse = rig["reverse"]
    log = rig["log"]
    config = rig["config"]
    cursor = rig["cursor"]

    wb, path = _open_ops(config)
    ws = _tasks_sheet(wb)
    row_idx = _row_index_by_task_id(ws, "t_sensitive")
    assert row_idx is not None
    ws.cell(row=row_idx, column=2, value="renamed sensitive")
    _save(wb, path)

    await reverse.run_cycle_now(OPS_WORKBOOK_NAME)

    updated = await _events_of_type(log, "task.updated", cursor)
    assert len(updated) == 1
    assert updated[0]["sensitivity"] == "sensitive"
    assert updated[0]["payload"]["task_id"] == "t_sensitive"
