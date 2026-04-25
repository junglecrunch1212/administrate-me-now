"""
Integration test: full xlsx round-trip (07c-β).

7 sqlite projections + forward daemon + reverse daemon (no watchdog
observer started; driven via ``run_cycle_now``). Seeds ~15 events across
Tasks, Commitments, Recurrences, accounts, and money_flows; regenerates
both workbooks; performs programmatic xlsx edits; runs reverse cycles;
asserts the full event sequence; regenerates again to prove the next
forward pass picks up the principal-authored deltas.
"""

from __future__ import annotations

import asyncio
import time
from pathlib import Path
from typing import Any, AsyncIterator

import pytest
from openpyxl import load_workbook

from adminme.daemons.xlsx_sync.reverse import XlsxReverseDaemon
from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"i" * 32


def _envelope(event_type: str, payload: dict[str, Any]) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, sid: str, timeout: float = 15.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        s = await bus.subscriber_status(sid)
        if s["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"{sid} stayed lagged: {s}")


async def _events_of_type(
    log: EventLog, event_type: str, after: str | None
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    async for e in log.read_since(after):
        if e["type"] == event_type:
            out.append(e)
    return out


def _row_index_by_id(ws: Any, value: Any, column: int = 1) -> int | None:
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[column - 1] == value:
            return i
    return None


@pytest.fixture
async def rig(tmp_path: Path) -> AsyncIterator[dict[str, Any]]:
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(TasksProjection())
    runner.register(CommitmentsProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    await runner.start()
    try:
        yield {
            "config": config,
            "log": log,
            "bus": bus,
            "runner": runner,
        }
    finally:
        await runner.stop()
        await log.close()


async def test_full_round_trip(rig: dict[str, Any]) -> None:
    config = rig["config"]
    log = rig["log"]
    bus = rig["bus"]
    runner = rig["runner"]

    # Seed ~15 events across the fixture surface area.
    await log.append(_envelope("party.created", {
        "party_id": "p1", "kind": "person",
        "display_name": "P1", "sort_name": "P1",
    }))
    await log.append(_envelope("party.created", {
        "party_id": "p2", "kind": "person",
        "display_name": "P2", "sort_name": "P2",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t1", "title": "first task",
    }))
    await log.append(_envelope("task.created", {
        "task_id": "t2", "title": "second task",
    }))
    await log.append(_envelope("commitment.proposed", {
        "commitment_id": "c1",
        "kind": "reply",
        "owed_by_member_id": "p1",
        "owed_to_party_id": "p2",
        "text_summary": "say hi",
        "confidence": 0.9,
        "strength": "confident",
    }))
    await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r1",
        "linked_kind": "household",
        "linked_id": "household",
        "kind": "maintenance",
        "rrule": "FREQ=MONTHLY",
        "next_occurrence": "2026-05-01T09:00:00Z",
    }))
    await log.append(_envelope("recurrence.added", {
        "recurrence_id": "r2",
        "linked_kind": "household",
        "linked_id": "household",
        "kind": "checkin",
        "rrule": "FREQ=WEEKLY",
        "next_occurrence": "2026-05-02T09:00:00Z",
    }))
    await log.append(_envelope("account.added", {
        "account_id": "a1", "display_name": "Checking",
        "organization_party_id": "p1", "kind": "bank",
        "status": "active", "attributes": {"last4": "1234"},
    }))
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "fp1", "amount_minor": 1000, "currency": "USD",
        "occurred_at": "2026-04-20T10:00:00Z",
        "kind": "paid", "linked_account": "a1",
        "source_adapter": "plaid",
    }))
    await log.append(_envelope("money_flow.recorded", {
        "flow_id": "fp2", "amount_minor": 2500, "currency": "USD",
        "occurred_at": "2026-04-21T10:00:00Z",
        "kind": "paid", "linked_account": "a1",
        "source_adapter": "plaid",
    }))
    last = await log.append(_envelope("money_flow.manually_added", {
        "flow_id": "fm1", "amount_minor": 4321, "currency": "USD",
        "occurred_at": "2026-04-22T10:00:00Z",
        "kind": "paid",
        "category": "groceries",
        "notes": "manual entry",
        "added_by_party_id": "p1",
    }))
    await bus.notify(last)
    for name in (
        "parties", "tasks", "commitments", "recurrences", "calendars",
        "places_assets_accounts", "money",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    forward = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.05)
    await forward.regenerate_now(OPS_WORKBOOK_NAME)
    await forward.regenerate_now(FINANCE_WORKBOOK_NAME)

    reverse = XlsxReverseDaemon(
        config,
        ctx,
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=2.0,
        delete_undo_window_s=0.1,
    )
    try:
        # ----- Ops workbook round trip -----
        ops_path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
        wb = load_workbook(str(ops_path))
        # Modify t1 title
        tasks_ws = wb["Tasks"]
        idx = _row_index_by_id(tasks_ws, "t1")
        assert idx is not None
        tasks_ws.cell(row=idx, column=2, value="renamed first task")
        # Append blank-id Task
        new_row = tasks_ws.max_row + 1
        tasks_ws.cell(row=new_row, column=2, value="appended blank-id task")
        # Attempt to delete a Recurrence
        rec_ws = wb["Recurrences"]
        rec_idx = _row_index_by_id(rec_ws, "r1")
        assert rec_idx is not None
        rec_ws.delete_rows(rec_idx)
        wb.save(str(ops_path))
        wb.close()

        cursor_pre_ops_cycle = await log.latest_event_id()
        await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
        # Wait past the undo window so any queued task.deleted resolves.
        await asyncio.sleep(0.3)

        # Recurrences DELETE drops INFO (descriptor.deletes_emit_event=None).
        rec_deleted = await _events_of_type(
            log, "recurrence.deleted", cursor_pre_ops_cycle
        )
        assert rec_deleted == []

        task_updated = await _events_of_type(
            log, "task.updated", cursor_pre_ops_cycle
        )
        assert len(task_updated) == 1
        assert task_updated[0]["payload"]["task_id"] == "t1"

        task_created = await _events_of_type(
            log, "task.created", cursor_pre_ops_cycle
        )
        # The original t1 / t2 already exist; only the appended blank-id row
        # produces a new task.created from this cycle.
        assert len(task_created) == 1
        new_task_id = task_created[0]["payload"]["task_id"]
        assert new_task_id.startswith("tsk_")
        assert task_created[0]["payload"]["title"] == "appended blank-id task"

        projected_ops = await _events_of_type(
            log, "xlsx.reverse_projected", cursor_pre_ops_cycle
        )
        assert len(projected_ops) == 1
        assert projected_ops[0]["payload"]["workbook_name"] == OPS_WORKBOOK_NAME
        # events_emitted contains the envelope ids of task.updated +
        # task.created (in some order).
        emitted_ids = projected_ops[0]["payload"]["events_emitted"]
        assert len(emitted_ids) >= 2
        assert task_updated[0]["event_id"] in emitted_ids
        assert task_created[0]["event_id"] in emitted_ids

        # Forward regenerate ops; principal-authored t1 rename + appended
        # task should now appear in the workbook.
        latest_after_emits = await log.latest_event_id()
        assert latest_after_emits is not None
        await bus.notify(latest_after_emits)
        await _wait_idle(bus, "projection:tasks")
        await forward.regenerate_now(OPS_WORKBOOK_NAME)
        wb = load_workbook(str(ops_path))
        tasks_ws = wb["Tasks"]
        rows_by_id = {
            r[0]: r for r in tasks_ws.iter_rows(min_row=2, values_only=True) if r[0]
        }
        assert "t1" in rows_by_id
        assert rows_by_id["t1"][1] == "renamed first task"
        assert new_task_id in rows_by_id
        assert rows_by_id[new_task_id][1] == "appended blank-id task"
        wb.close()

        # ----- Finance workbook round trip -----
        finance_path = config.xlsx_workbooks_dir / FINANCE_WORKBOOK_NAME
        wb = load_workbook(str(finance_path))
        ws = wb["Raw Data"]
        # Append manual row
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value="")  # mint
        ws.cell(row=new_row, column=2, value="2026-04-25")
        ws.cell(row=new_row, column=6, value=20.00)
        ws.cell(row=new_row, column=9, value="dining")
        ws.cell(row=new_row, column=11, value=True)
        # Append non-manual row (should drop)
        new_row += 1
        ws.cell(row=new_row, column=1, value="plaid_x")
        ws.cell(row=new_row, column=2, value="2026-04-26")
        ws.cell(row=new_row, column=6, value=99.99)
        ws.cell(row=new_row, column=11, value=False)
        # Delete the seeded manual row fm1
        idx = _row_index_by_id(ws, "fm1")
        assert idx is not None
        ws.delete_rows(idx)
        wb.save(str(finance_path))
        wb.close()

        cursor_pre_finance_cycle = await log.latest_event_id()
        await reverse.run_cycle_now(FINANCE_WORKBOOK_NAME)
        await asyncio.sleep(0.3)

        added = await _events_of_type(
            log, "money_flow.manually_added", cursor_pre_finance_cycle
        )
        # Exactly one manual ADD landed (Plaid add was dropped).
        assert len(added) == 1
        new_flow_id = added[0]["payload"]["flow_id"]
        assert new_flow_id.startswith("flow_")

        deleted = await _events_of_type(
            log, "money_flow.manually_deleted", cursor_pre_finance_cycle
        )
        assert len(deleted) == 1
        assert deleted[0]["payload"]["flow_id"] == "fm1"

        projected_fin = await _events_of_type(
            log, "xlsx.reverse_projected", cursor_pre_finance_cycle
        )
        assert len(projected_fin) == 1
        assert (
            projected_fin[0]["payload"]["workbook_name"] == FINANCE_WORKBOOK_NAME
        )
    finally:
        await reverse.stop()


# ---------------------------------------------------------------------------
# UT-7 closure (08b): principal_member_id attribution
# ---------------------------------------------------------------------------


async def test_ut7_reverse_emits_attribute_detected_principal(
    rig: dict[str, Any],
) -> None:
    """UT-7 closure (08b): when the reverse daemon is wired with a
    principal_member_id_resolver that returns a specific member id,
    every domain event emitted by that cycle carries
    ``actor_identity == <detected principal>`` (not the literal
    ``"xlsx_reverse"``)."""
    config = rig["config"]
    log = rig["log"]
    bus = rig["bus"]
    runner = rig["runner"]

    # Seed a single party + an existing task so the cycle has both an ADD
    # path and an UPDATE path to exercise.
    await log.append(_envelope("party.created", {
        "party_id": "principal_a", "kind": "person",
        "display_name": "principal_a", "sort_name": "principal_a",
    }))
    seed = await log.append(_envelope("task.created", {
        "task_id": "ut7_t1", "title": "ut7 baseline task",
    }))
    await bus.notify(seed)
    for name in ("parties", "tasks"):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    forward = XlsxWorkbooksProjection(config, ctx, event_log=log, debounce_s=0.05)
    await forward.regenerate_now(OPS_WORKBOOK_NAME)

    detected_principal = "principal_a"
    reverse = XlsxReverseDaemon(
        config,
        ctx,
        event_log=log,
        flush_wait_s=0.05,
        forward_lock_timeout_s=2.0,
        delete_undo_window_s=0.1,
        principal_member_id_resolver=lambda _wb: detected_principal,
    )
    try:
        # Edit the existing task title + append a brand-new task. Both
        # emits must attribute to detected_principal.
        ops_path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
        wb = load_workbook(str(ops_path))
        tasks_ws = wb["Tasks"]
        idx = _row_index_by_id(tasks_ws, "ut7_t1")
        assert idx is not None
        tasks_ws.cell(row=idx, column=2, value="ut7 renamed by principal")
        new_row = tasks_ws.max_row + 1
        tasks_ws.cell(row=new_row, column=2, value="ut7 appended by principal")
        wb.save(str(ops_path))
        wb.close()

        cursor_pre = await log.latest_event_id()
        await reverse.run_cycle_now(OPS_WORKBOOK_NAME)
        await asyncio.sleep(0.3)

        # task.created — appended row
        created = await _events_of_type(log, "task.created", cursor_pre)
        assert len(created) == 1
        assert created[0]["actor_identity"] == detected_principal
        assert created[0]["payload"]["title"] == "ut7 appended by principal"

        # task.updated — title rename
        updated = await _events_of_type(log, "task.updated", cursor_pre)
        assert len(updated) == 1
        assert updated[0]["actor_identity"] == detected_principal
        assert updated[0]["payload"]["task_id"] == "ut7_t1"
        # The updated_by_party_id payload field also reflects the principal,
        # not the literal placeholder.
        assert updated[0]["payload"]["updated_by_party_id"] == detected_principal

        # The terminal cycle event stays system-attributed.
        projected = await _events_of_type(
            log, "xlsx.reverse_projected", cursor_pre
        )
        assert len(projected) == 1
        assert projected[0]["actor_identity"] == "xlsx_reverse"
    finally:
        await reverse.stop()
