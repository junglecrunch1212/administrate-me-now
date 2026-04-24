"""
Integration test: xlsx forward end-to-end.

Per phase 07b prompt Commit 4. Populates ~50 events across trigger types,
registers all 10 projections plus XlsxWorkbooksProjection, verifies both
workbooks build correctly, and asserts the §2.2 invariant that the xlsx
projection emits ONLY ``xlsx.regenerated`` system events.
"""

from __future__ import annotations

import asyncio
import time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from adminme.events.bus import EventBus
from adminme.events.envelope import EventEnvelope
from adminme.events.log import EventLog
from adminme.lib.instance_config import load_instance_config
from adminme.projections.artifacts import ArtifactsProjection
from adminme.projections.calendars import CalendarsProjection
from adminme.projections.commitments import CommitmentsProjection
from adminme.projections.interactions import InteractionsProjection
from adminme.projections.money import MoneyProjection
from adminme.projections.parties import PartiesProjection
from adminme.projections.places_assets_accounts import PlacesAssetsAccountsProjection
from adminme.projections.recurrences import RecurrencesProjection
from adminme.projections.runner import ProjectionRunner
from adminme.projections.tasks import TasksProjection
from adminme.projections.vector_search import VectorSearchProjection
from adminme.projections.xlsx_workbooks import (
    FINANCE_WORKBOOK_NAME,
    OPS_WORKBOOK_NAME,
    XlsxWorkbooksProjection,
)
from adminme.projections.xlsx_workbooks.query_context import XlsxQueryContext

TEST_KEY = b"y" * 32


def _envelope(event_type: str, payload: dict[str, Any]) -> EventEnvelope:
    return EventEnvelope(
        event_at_ms=int(time.time() * 1000),
        tenant_id="tenant-a",
        type=event_type,
        schema_version=1,
        occurred_at=EventEnvelope.now_utc_iso(),
        source_adapter="test",
        source_account_id="test-acct",
        owner_scope="shared:household",
        visibility_scope="shared:household",
        sensitivity="normal",
        payload=payload,
    )


async def _wait_idle(bus: EventBus, subscriber_id: str, timeout: float = 15.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        status = await bus.subscriber_status(subscriber_id)
        if status["lag_count"] == 0:
            return
        await asyncio.sleep(0.02)
    raise AssertionError(f"subscriber {subscriber_id} stayed lagged: {status}")


@pytest.fixture
async def rig(tmp_path: Path):
    instance_dir = tmp_path / "instance"
    instance_dir.mkdir()
    (instance_dir / "config").mkdir()
    (instance_dir / "config" / "instance.yaml").write_text("tenant_id: tenant-a\n")
    config = load_instance_config(instance_dir)
    log = EventLog(config, TEST_KEY)
    bus = EventBus(log, config.bus_checkpoint_path)
    runner = ProjectionRunner(bus, log, config, encryption_key=TEST_KEY)
    runner.register(PartiesProjection())
    runner.register(InteractionsProjection())
    runner.register(ArtifactsProjection())
    runner.register(CommitmentsProjection())
    runner.register(TasksProjection())
    runner.register(RecurrencesProjection())
    runner.register(CalendarsProjection())
    runner.register(PlacesAssetsAccountsProjection())
    runner.register(MoneyProjection())
    runner.register(VectorSearchProjection())
    await runner.start()
    try:
        yield {"config": config, "log": log, "bus": bus, "runner": runner}
    finally:
        await runner.stop()
        await log.close()


async def test_xlsx_forward_end_to_end(rig: dict[str, Any]) -> None:
    config = rig["config"]
    log = rig["log"]
    bus = rig["bus"]
    runner = rig["runner"]

    # 5 persons + 2 orgs + 1 household
    for i in range(5):
        await log.append(_envelope("party.created", {
            "party_id": f"p{i}", "kind": "person",
            "display_name": f"P{i}", "sort_name": f"P{i}",
        }))
    for i in range(2):
        await log.append(_envelope("party.created", {
            "party_id": f"org{i}", "kind": "organization",
            "display_name": f"Org{i}", "sort_name": f"Org{i}",
        }))
    await log.append(_envelope("party.created", {
        "party_id": "hh", "kind": "household",
        "display_name": "HH", "sort_name": "HH",
    }))

    # 10 tasks
    for i in range(10):
        await log.append(_envelope("task.created", {
            "task_id": f"t{i}", "title": f"Task {i}",
            "owner_member_id": f"p{i % 5}",
        }))

    # 5 commitments
    for i in range(5):
        await log.append(_envelope("commitment.proposed", {
            "commitment_id": f"c{i}",
            "kind": "reply",
            "owed_by_member_id": f"p{i}",
            "owed_to_party_id": f"org{i % 2}",
            "text_summary": f"Commit {i}",
            "confidence": 0.8,
            "strength": "confident",
        }))

    # 3 recurrences
    for i in range(3):
        await log.append(_envelope("recurrence.added", {
            "recurrence_id": f"r{i}",
            "linked_kind": "household",
            "linked_id": "hh",
            "kind": "maintenance",
            "rrule": "FREQ=MONTHLY",
            "next_occurrence": f"2026-05-0{i+1}T09:00:00Z",
        }))

    # 2 accounts + 5 money flows
    for i in range(2):
        await log.append(_envelope("account.added", {
            "account_id": f"a{i}", "display_name": f"Acc{i}",
            "organization_party_id": f"org{i % 2}",
            "kind": "bank",
            "status": "active",
            "attributes": {"last4": f"{1000 + i}"},
        }))
    for i in range(5):
        await log.append(_envelope("money_flow.recorded", {
            "flow_id": f"f{i}",
            "amount_minor": 1000 + i * 100,
            "currency": "USD",
            "occurred_at": f"2026-04-{20+i:02d}T10:00:00Z",
            "kind": "paid",
            "linked_account": f"a{i % 2}",
            "source_adapter": "plaid",
        }))

    # 2 places + 2 assets
    for i in range(2):
        await log.append(_envelope("place.added", {
            "place_id": f"pl{i}", "display_name": f"Place{i}",
            "kind": "home", "address_json": {},
        }))
        await log.append(_envelope("asset.added", {
            "asset_id": f"as{i}", "display_name": f"Asset{i}",
            "kind": "vehicle",
        }))

    # 2 calendar events
    for i in range(2):
        await log.append(_envelope("calendar.event_added", {
            "source": "google",
            "external_event_id": f"ext{i}",
            "calendar_id": "cal1",
            "summary": f"Event {i}",
            "start": f"2026-04-25T{10+i:02d}:00:00Z",
            "end": f"2026-04-25T{11+i:02d}:00:00Z",
        }))

    latest = await log.latest_event_id()
    assert latest is not None
    await bus.notify(latest)
    for name in (
        "parties", "interactions", "artifacts", "commitments", "tasks",
        "recurrences", "calendars", "places_assets_accounts", "money",
        "vector_search",
    ):
        await _wait_idle(bus, f"projection:{name}")

    ctx = XlsxQueryContext(
        parties_conn=runner.connection("parties"),
        tasks_conn=runner.connection("tasks"),
        commitments_conn=runner.connection("commitments"),
        recurrences_conn=runner.connection("recurrences"),
        calendars_conn=runner.connection("calendars"),
        places_assets_accounts_conn=runner.connection("places_assets_accounts"),
        money_conn=runner.connection("money"),
    )
    xlsx_projection = XlsxWorkbooksProjection(
        config, ctx, event_log=log, debounce_s=0.05
    )

    # Capture the cursor before any xlsx.regenerated appends.
    cursor_before_regens = await log.latest_event_id()
    assert cursor_before_regens is not None

    # Regenerate both workbooks; bypass the debounce.
    await xlsx_projection.regenerate_now(OPS_WORKBOOK_NAME)
    await xlsx_projection.regenerate_now(FINANCE_WORKBOOK_NAME)

    ops_path = config.xlsx_workbooks_dir / OPS_WORKBOOK_NAME
    finance_path = config.xlsx_workbooks_dir / FINANCE_WORKBOOK_NAME
    assert ops_path.exists()
    assert finance_path.exists()

    wb_ops = load_workbook(str(ops_path))
    assert set(wb_ops.sheetnames) == {
        "Tasks", "Recurrences", "Commitments", "People", "Metadata"
    }
    assert wb_ops["Tasks"].max_row == 11  # 10 tasks + header
    assert wb_ops["Commitments"].max_row == 6
    assert wb_ops["Recurrences"].max_row == 4

    wb_fin = load_workbook(str(finance_path))
    assert set(wb_fin.sheetnames) == {"Raw Data", "Accounts", "Metadata"}
    assert wb_fin["Raw Data"].max_row == 6  # 5 flows + header
    assert wb_fin["Accounts"].max_row == 3

    # Exactly two xlsx.regenerated events after the cursor
    xlsx_events: list[dict[str, Any]] = []
    async for e in log.read_since(cursor_before_regens):
        xlsx_events.append(e)
    xlsx_types = [e["type"] for e in xlsx_events]
    assert xlsx_types == ["xlsx.regenerated", "xlsx.regenerated"]

    # §2.2 audit: no event after the cursor was sourced by
    # xlsx_workbooks EXCEPT xlsx.regenerated.
    for e in xlsx_events:
        if e["source_adapter"] == "xlsx_workbooks":
            assert e["type"] == "xlsx.regenerated", (
                f"xlsx projection emitted non-system event: {e['type']}"
            )

    # Third regenerate with no intervening events → semantically equal
    # (compared cell-by-cell on the Tasks sheet).
    def _data_rows(ws: Any) -> list[tuple]:
        return [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]

    before = _data_rows(wb_ops["Tasks"])
    await xlsx_projection.regenerate_now(OPS_WORKBOOK_NAME)
    wb_ops2 = load_workbook(str(ops_path))
    after = _data_rows(wb_ops2["Tasks"])
    assert before == after
